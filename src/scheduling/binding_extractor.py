"""
绑定人员提取程序
从绑定集合表和总表中提取各小组的绑定人员信息

输入：绑定集合表.xlsx、总表.xlsx
输出：绑定人员明细表.xlsx

功能：筛选出绑定类型为couple和family的绑定集合，查找其在总表中的小组号，
     生成包含小组号和绑定人员姓名的汇总表
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.append(str(project_root))

from src.utils.logger_factory import get_logger
from src.utils._excel_handler import ExcelHandler
from config.loader import CONFIG


class BindingExtractor:
    """绑定人员提取器"""

    def __init__(self):
        self.logger = get_logger(__file__)
        self.handler = ExcelHandler()

        # 配置路径
        self.output_dir = CONFIG.get('paths.output_dir')
        self.scheduling_prep_dir = CONFIG.get('paths.scheduling_prep_dir')

    def extract_binding_members(self) -> Dict[str, Any]:
        """提取绑定人员信息"""
        self.logger.info("开始提取绑定人员信息")

        results = {
            'output_file': None,
            'binding_count': 0,
            'statistics': {},
            'errors': [],
            'warnings': []
        }

        try:
            # 步骤1：读取绑定集合表
            binding_sets_df = self._read_binding_sets()

            # 步骤2：筛选couple和family类型的绑定集合
            filtered_bindings = self._filter_binding_types(
                binding_sets_df, ['couple', 'family']
            )

            if filtered_bindings.empty:
                self.logger.warning("未找到couple或family类型的绑定集合")
                results['warnings'].append("未找到couple或family类型的绑定集合")
                return results

            # 步骤3：读取总表
            master_schedule_df = self._read_master_schedule()

            # 步骤4：查找每个绑定集合成员的小组号
            binding_groups = self._find_binding_groups(
                filtered_bindings, master_schedule_df
            )

            # 步骤5：生成输出数据
            output_data = self._prepare_output_data(binding_groups)

            # 步骤6：保存到Excel
            output_file = self._save_to_excel(output_data)

            # 步骤7：美化Excel格式
            self._format_excel(output_file)

            # 步骤8：统计信息
            statistics = self._calculate_statistics(output_data)

            results.update({
                'output_file': output_file,
                'binding_count': len(output_data),
                'statistics': statistics
            })

            self.logger.info(f"绑定人员提取完成：共提取 {len(output_data)} 个绑定集合")

        except Exception as e:
            self.logger.error(f"绑定人员提取失败: {str(e)}")
            results['errors'].append(str(e))
            raise

        return results

    def _read_binding_sets(self) -> pd.DataFrame:
        """读取绑定集合表"""
        binding_file = os.path.join(
            self.scheduling_prep_dir,
            CONFIG.get('files.binding_sets')
        )

        if not os.path.exists(binding_file):
            raise FileNotFoundError(f"绑定集合表文件不存在: {binding_file}")

        df = self.handler.read_excel(binding_file)
        self.logger.info(f"读取绑定集合表: {len(df)} 行")
        return df

    def _read_master_schedule(self) -> pd.DataFrame:
        """读取总表"""
        master_file = os.path.join(
            self.output_dir,
            CONFIG.get('files.master_schedule')
        )

        if not os.path.exists(master_file):
            raise FileNotFoundError(f"总表文件不存在: {master_file}")

        df = self.handler.read_excel(master_file)
        self.logger.info(f"读取总表: {len(df)} 行")
        return df

    def _filter_binding_types(
        self,
        binding_df: pd.DataFrame,
        target_types: List[str]
    ) -> pd.DataFrame:
        """筛选指定类型的绑定集合"""
        if '绑定类型' not in binding_df.columns:
            raise ValueError("绑定集合表中缺少'绑定类型'列")

        filtered = binding_df[binding_df['绑定类型'].isin(target_types)].copy()
        self.logger.info(
            f"筛选出 {target_types} 类型的绑定记录: {len(filtered)} 行"
        )
        return filtered

    def _find_binding_groups(
        self,
        binding_df: pd.DataFrame,
        master_df: pd.DataFrame
    ) -> List[Dict[str, Any]]:
        """查找每个绑定集合成员的小组号"""
        self.logger.info("开始查找绑定集合成员的小组号")

        # 检查必要列
        required_binding_cols = ['绑定集合ID', '成员学号', '成员姓名', '绑定类型']
        for col in required_binding_cols:
            if col not in binding_df.columns:
                raise ValueError(f"绑定集合表中缺少'{col}'列")

        if '小组号' not in master_df.columns:
            raise ValueError("总表中缺少'小组号'列")
        if '学号' not in master_df.columns:
            raise ValueError("总表中缺少'学号'列")

        # 创建学号到小组号的映射
        student_to_group = {}
        for _, row in master_df.iterrows():
            student_id = str(row['学号']).strip()
            group_id = row['小组号']
            student_to_group[student_id] = group_id

        # 按绑定集合ID分组
        binding_groups = []
        grouped = binding_df.groupby('绑定集合ID')

        for binding_id, group_df in grouped:
            # 获取绑定类型（应该是相同的）
            binding_type = group_df.iloc[0]['绑定类型']

            # 收集成员信息
            members = []
            group_ids = set()

            for _, row in group_df.iterrows():
                student_id = str(row['成员学号']).strip()
                member_name = str(row['成员姓名']).strip()

                # 查找小组号
                group_id = student_to_group.get(student_id)

                if group_id is not None:
                    group_ids.add(group_id)
                    members.append({
                        'student_id': student_id,
                        'name': member_name,
                        'group_id': group_id
                    })
                else:
                    self.logger.warning(
                        f"绑定集合 {binding_id} 的成员 {member_name}({student_id}) "
                        f"未在总表中找到"
                    )

            # 检查所有成员是否在同一小组
            if len(group_ids) == 0:
                self.logger.warning(
                    f"绑定集合 {binding_id} 的所有成员都未在总表中找到"
                )
                continue
            elif len(group_ids) > 1:
                self.logger.warning(
                    f"绑定集合 {binding_id} 的成员分布在不同小组: {group_ids}"
                )

            # 使用第一个成员的小组号（正常情况下应该都相同）
            if members:
                binding_groups.append({
                    'binding_id': binding_id,
                    'binding_type': binding_type,
                    'group_id': members[0]['group_id'],
                    'members': members
                })

        self.logger.info(f"成功匹配 {len(binding_groups)} 个绑定集合")
        return binding_groups

    def _prepare_output_data(
        self,
        binding_groups: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """准备输出数据"""
        output_data = []

        for binding_group in binding_groups:
            group_id = binding_group['group_id']
            members = binding_group['members']
            binding_type = binding_group['binding_type']

            # 准备输出行
            row = {
                '小组号': group_id,
                '绑定类型': binding_type
            }

            # 填充人员姓名（最多3个）
            for i in range(3):
                if i < len(members):
                    row[f'人员{self._number_to_chinese(i+1)}'] = members[i]['name']
                else:
                    row[f'人员{self._number_to_chinese(i+1)}'] = ''

            output_data.append(row)

        # 按小组号排序
        output_data.sort(key=lambda x: x['小组号'])
        return output_data

    def _number_to_chinese(self, num: int) -> str:
        """数字转中文"""
        chinese_numbers = {1: '一', 2: '二', 3: '三'}
        return chinese_numbers.get(num, str(num))

    def _save_to_excel(self, output_data: List[Dict[str, Any]]) -> str:
        """保存到Excel文件"""
        output_file = os.path.join(self.output_dir, '绑定人员明细表.xlsx')

        # 转换为DataFrame
        df = pd.DataFrame(output_data)

        # 调整列顺序
        columns_order = ['小组号', '绑定类型', '人员一', '人员二', '人员三']
        df = df[columns_order]

        # 保存到Excel
        self.handler.write_excel(df, output_file)
        self.logger.info(f"保存绑定人员明细表: {output_file}")

        return output_file

    def _format_excel(self, file_path: str):
        """美化Excel格式"""
        try:
            # 加载工作簿
            wb = load_workbook(file_path)
            ws = wb.active

            # 设置标题行格式
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )

            # 应用标题行样式
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill

            # 设置列宽
            column_widths = {
                'A': 12,  # 小组号
                'B': 15,  # 绑定类型
                'C': 15,  # 人员一
                'D': 15,  # 人员二
                'E': 15   # 人员三
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # 设置数据对齐
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

            # 冻结首行
            ws.freeze_panes = 'A2'

            # 保存工作簿
            wb.save(file_path)
            self.logger.info(f"Excel格式美化完成: {file_path}")

        except Exception as e:
            self.logger.warning(f"美化Excel文件失败: {str(e)}")

    def _calculate_statistics(
        self,
        output_data: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """计算统计信息"""
        stats = {
            'total_bindings': len(output_data),
            'by_type': {},
            'by_member_count': {},
            'groups_with_bindings': set()
        }

        for row in output_data:
            binding_type = row['绑定类型']
            group_id = row['小组号']

            # 按类型统计
            if binding_type not in stats['by_type']:
                stats['by_type'][binding_type] = 0
            stats['by_type'][binding_type] += 1

            # 统计成员数量
            member_count = sum(
                1 for i in range(1, 4)
                if row.get(f'人员{self._number_to_chinese(i)}', '')
            )

            if member_count not in stats['by_member_count']:
                stats['by_member_count'][member_count] = 0
            stats['by_member_count'][member_count] += 1

            # 记录包含绑定的小组
            stats['groups_with_bindings'].add(group_id)

        stats['groups_with_bindings'] = len(stats['groups_with_bindings'])

        return stats


def main():
    """主函数"""
    print("=" * 60)
    print("绑定人员提取程序")
    print("=" * 60)

    extractor = BindingExtractor()

    try:
        results = extractor.extract_binding_members()

        if results['errors']:
            print("\n处理过程中发生错误：")
            for error in results['errors']:
                print(f"  - {error}")
            return False

        if results['warnings']:
            print("\n警告信息：")
            for warning in results['warnings']:
                print(f"  - {warning}")

        print(f"\n✓ 绑定人员提取完成！")
        print(f"  - 输出文件: {results['output_file']}")
        print(f"  - 绑定集合数: {results['binding_count']}")

        if results['statistics']:
            stats = results['statistics']
            print(f"\n统计信息：")
            print(f"  - 总绑定集合数: {stats['total_bindings']}")
            print(f"  - 涉及小组数: {stats['groups_with_bindings']}")

            print(f"\n  按绑定类型统计：")
            for binding_type, count in stats['by_type'].items():
                print(f"    - {binding_type}: {count}")

            print(f"\n  按成员数量统计：")
            for member_count, count in stats['by_member_count'].items():
                print(f"    - {member_count}人: {count}个绑定集合")

        return True

    except Exception as e:
        print(f"\n✗ 程序执行失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
