"""
总表拆分程序
程序七：总表拆分和整合

输入：总表Excel文件
输出：多个小组拆分后的Excel文件

功能一（总表拆分）：读取总表Excel文件，按照"小组号"字段进行拆分，生成多个新的Excel文件
功能二（表格整合）：将重要表格整合到一个Excel文件中，生成"大总表"
"""

import os
import sys
import json
from pathlib import Path
from typing import Dict, List, Tuple, Any
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.append(str(project_root))

from src.utils.logger_factory import get_logger
from src.utils._excel_handler import ExcelHandler
from config.loader import CONFIG


class Finalizer:
    """最终处理器"""

    def __init__(self):
        self.logger = get_logger(__file__)
        self.handler = ExcelHandler()

        # 配置路径
        self.output_dir = CONFIG.get('paths.output_dir')
        self.scheduling_prep_dir = CONFIG.get('paths.scheduling_prep_dir')
        self.groups_output_dir = CONFIG.get('paths.groups_output_dir')

        # 确保目录存在
        os.makedirs(self.groups_output_dir, exist_ok=True)

    def run_finalization(self) -> Dict[str, Any]:
        """执行最终处理流程"""
        self.logger.info("开始执行最终处理流程")

        results = {
            'split_files': {},
            'integrated_file': None,
            'statistics': {},
            'errors': [],
            'warnings': []
        }

        try:
            # 步骤1：读取总表
            master_schedule_df = self._read_master_schedule()

            # 步骤2：拆分总表为小组文件
            split_files = self._split_master_schedule(master_schedule_df)

            # 步骤3：生成整合大总表
            integrated_file = self._generate_integrated_schedule()

            # 步骤4：统计信息
            statistics = self._calculate_finalization_statistics(split_files, integrated_file)

            results.update({
                'split_files': split_files,
                'integrated_file': integrated_file,
                'statistics': statistics
            })

            self.logger.info("最终处理流程执行完成")

        except Exception as e:
            self.logger.error(f"最终处理流程执行失败: {str(e)}")
            results['errors'].append(str(e))

        return results

    def _read_master_schedule(self) -> pd.DataFrame:
        """读取总表"""
        master_file = os.path.join(self.output_dir, CONFIG.get('files.master_schedule'))

        if not os.path.exists(master_file):
            raise FileNotFoundError(f"总表文件不存在: {master_file}")

        df = self.handler.read_excel(master_file)
        self.logger.info(f"读取总表: {len(df)} 行")
        return df

    def _split_master_schedule(self, master_df: pd.DataFrame) -> Dict[int, str]:
        """拆分总表为小组文件"""
        self.logger.info("开始拆分总表为小组文件")

        if '小组号' not in master_df.columns:
            raise ValueError("总表中缺少'小组号'列")

        split_files = {}

        # 按小组号分组
        grouped = master_df.groupby('小组号')

        for group_number, group_df in grouped:
            try:
                # 移除敏感列（证件类型、证件号）
                columns_to_remove = ['证件类型', '证件号']
                columns_to_keep = [col for col in group_df.columns if col not in columns_to_remove]

                # 创建小组数据
                group_data = group_df[columns_to_keep].copy()

                # 保存小组文件
                output_file = os.path.join(self.groups_output_dir, f"{group_number}.xlsx")
                self.handler.write_excel(group_data, output_file)

                # 美化小组文件格式
                self._format_group_file(output_file)

                split_files[group_number] = output_file
                self.logger.info(f"生成小组 {group_number} 文件: {len(group_data)} 行")

            except Exception as e:
                self.logger.error(f"生成小组 {group_number} 文件失败: {str(e)}")
                continue

        self.logger.info(f"总表拆分完成：生成 {len(split_files)} 个小组文件")
        return split_files

    def _format_group_file(self, file_path: str):
        """美化小组文件格式"""
        try:
            # 加载工作簿
            wb = load_workbook(file_path)
            ws = wb.active

            # 设置标题行格式
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # 设置标题行样式
            for cell in ws[1]:  # 第一行
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill

            # 设置列宽
            column_widths = {
                'A': 10,  # 小组号
                'B': 15,  # 岗位名称
                'C': 20,  # 岗位简介
                'D': 15,  # 学号
                'E': 12,  # 姓名
                'F': 8,   # 性别
                'G': 15,  # 学院
                'H': 15,  # 手机号
                'I': 25,  # 邮箱
                'J': 12,  # 宿舍楼栋
                'K': 10,  # 衣服尺码
                'L': 12,  # 志愿者类型
                'M': 8    # 是否组长
            }

            for col, width in column_widths.items():
                if col in [cell.column_letter for cell in ws[1]]:
                    ws.column_dimensions[col].width = width

            # 设置数据对齐
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # 冻结首行
            ws.freeze_panes = 'A2'

            # 保存工作簿
            wb.save(file_path)

        except Exception as e:
            self.logger.warning(f"美化小组文件 {file_path} 失败: {str(e)}")

    def _generate_integrated_schedule(self) -> str:
        """生成整合大总表"""
        self.logger.info("开始生成整合大总表")

        # 创建Excel写入器
        integrated_file = os.path.join(self.output_dir, CONFIG.get('files.integrated_schedule'))
        master_file = os.path.join(self.output_dir, CONFIG.get('files.master_schedule'))

        # 直接将原总表另存为大总表
        if not os.path.exists(master_file):
            self.logger.error(f"原总表文件不存在: {master_file}")
            return ""

        try:
            # 使用shutil直接复制文件，保留所有格式
            import shutil
            shutil.copy2(master_file, integrated_file)
            self.logger.info(f"已将原总表复制为: {os.path.basename(integrated_file)}")

        except Exception as e:
            self.logger.error(f"复制总表文件失败: {str(e)}")
            return ""

        # 在大总表基础上添加其他sheet
        self._add_additional_sheets(integrated_file)

        self.logger.info(f"整合大总表已生成: {integrated_file}")
        return integrated_file

    def _generate_color_table(self) -> pd.DataFrame:
        """生成颜色对照表"""
        self.logger.info("生成颜色对照表")

        colors = CONFIG.get('colors', {})

        # 身份/属性颜色
        color_data = [
            {'类型': '身份/属性', '名称': '组长', '颜色代码': colors.get('leader', 'FFFF00'), '颜色说明': '黄色'},
            {'类型': '身份/属性', '名称': '小闪电', '颜色代码': colors.get('lightning', '00FF00'), '颜色说明': '绿色'},
            {'类型': '身份/属性', '名称': '摄影', '颜色代码': colors.get('photography', 'E6E6FA'), '颜色说明': '淡紫色'},
            {'类型': '身份/属性', '名称': '情侣', '颜色代码': colors.get('couple', 'FFB6C1'), '颜色说明': '粉色'},
            {'类型': '身份/属性', '名称': '内部志愿者', '颜色代码': colors.get('internal', 'FFA500'), '颜色说明': '橙色'},
            {'类型': '身份/属性', '名称': '家属志愿者', '颜色代码': colors.get('family', '87CEEB'), '颜色说明': '天蓝色'},
            {'类型': '身份/属性', '名称': '普通志愿者', '颜色代码': colors.get('default', 'FFFFFF'), '颜色说明': '白色（无背景色）'},
        ]

        # 添加分隔行
        color_data.append({'类型': '', '名称': '', '颜色代码': '', '颜色说明': ''})

        # 从metadata.json读取团体颜色信息
        try:
            metadata_file = os.path.join(self.scheduling_prep_dir, CONFIG.get('files.metadata'))
            if os.path.exists(metadata_file):
                with open(metadata_file, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)

                group_colors = metadata.get('group_color_mapping', {})
                self.logger.info(f"从metadata读取到 {len(group_colors)} 个团体颜色映射")

                if group_colors:
                    # 按团体名称排序
                    for group_name in sorted(group_colors.keys()):
                        color_code = group_colors[group_name]
                        # 生成颜色描述
                        color_desc = self._get_color_description(color_code)
                        color_data.append({
                            '类型': '团体志愿者',
                            '名称': group_name,
                            '颜色代码': color_code,
                            '颜色说明': color_desc
                        })
                else:
                    self.logger.warning("metadata.json中未找到团体颜色映射")
            else:
                self.logger.warning(f"metadata.json文件不存在: {metadata_file}")

        except Exception as e:
            self.logger.error(f"读取团体颜色信息失败: {str(e)}")

        return pd.DataFrame(color_data)

    def _get_color_description(self, color_code: str) -> str:
        """根据颜色代码生成颜色描述"""
        # 常见颜色映射
        color_map = {
            '98FB98': '浅绿色',
            'DDA0DD': '梅红色',
            'F0E68C': '卡其色',
            'ADD8E6': '浅蓝色',
            'F5DEB3': '小麦色',
            'FFDAB9': '桃色',
            'E0FFFF': '浅青色',
            'FAFAD2': '浅黄色',
            'D3D3D3': '浅灰色',
            'FFE4B5': '莫卡辛色',
            'FFFACD': '柠檬绸色',
            'F0FFF0': '蜜露色',
            'FFC0CB': '粉色',
            '87CEEB': '天蓝色',
            'FFA500': '橙色',
            'E6E6FA': '淡紫色',
            'FFFF00': '黄色',
            '00FF00': '绿色',
            'FFB6C1': '粉色'
        }

        return color_map.get(color_code.upper(), f'自定义颜色({color_code})')

    def _add_additional_sheets(self, integrated_file: str):
        """在已有的大总表文件基础上添加其他sheet"""
        try:
            self.logger.info("开始在大总表基础上添加其他sheet")

            # 生成颜色对照表
            color_table_df = self._generate_color_table()

            # 打开已有的大总表文件
            from openpyxl import load_workbook
            wb = load_workbook(integrated_file)

            # 添加小组信息表
            group_info_file = os.path.join(self.scheduling_prep_dir, CONFIG.get('files.group_info'))
            if os.path.exists(group_info_file):
                group_info_df = self.handler.read_excel(group_info_file)
                ws_group = wb.create_sheet(title='小组信息表')

                # 写入标题
                headers = list(group_info_df.columns)
                for col_idx, header in enumerate(headers, 1):
                    header_cell = ws_group.cell(row=1, column=col_idx, value=header)
                    # 设置标题格式
                    header_cell.font = Font(bold=True, color="FFFFFF")
                    header_cell.alignment = Alignment(horizontal="center", vertical="center")
                    header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

                # 写入数据
                for row_idx, (_, row) in enumerate(group_info_df.iterrows(), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws_group.cell(row=row_idx, column=col_idx, value=value)
                        # 设置数据居中对齐
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                        # 为小组号列设置特殊背景色
                        if col_idx == 1:  # 第一列是小组号
                            cell.fill = PatternFill(
                                start_color="E6E6FA",
                                end_color="E6E6FA",
                                fill_type="solid"
                            )
                            cell.font = Font(bold=True)

                # 设置列宽
                column_widths = {
                    'A': 12,  # 小组号
                    'B': 25,  # 岗位名称
                    'C': 15,  # 需求人数
                    'D': 15,  # 实际人数
                    'E': 20,  # 组长学号
                    'F': 15,  # 组长姓名
                    'G': 15,  # 小闪电学号
                    'H': 15,  # 小闪电姓名
                    'I': 15,  # 摄影学号
                    'J': 15,  # 摄影姓名
                    'K': 20,  # 工作地点
                    'L': 30,  # 岗位简介
                }
                for col_letter, width in column_widths.items():
                    ws_group.column_dimensions[col_letter].width = width

                # 冻结首行
                ws_group.freeze_panes = 'A2'

                self.logger.info(f"小组信息表添加完成: {len(group_info_df)} 行（已美化格式）")

            # 添加储备志愿者表
            backup_file = os.path.join(self.scheduling_prep_dir, CONFIG.get('files.backup_volunteers'))
            if os.path.exists(backup_file):
                backup_df = self.handler.read_excel(backup_file)
                ws_backup = wb.create_sheet(title='储备志愿者表')

                # 定义需要保留的关键字段
                required_keywords = [
                    '学号', '姓名', '姓名拼音', '性别', '证件类型', '证件号', '出生日期',
                    '学院', '身高', '邮件', '手机号', 'QQ号', '微信号', '政治面貌',
                    '第几次参加马拉松志愿者', '校区', '宿舍楼栋', '衣服尺码'
                ]

                # 筛选包含关键字的列
                filtered_columns = []
                for col in backup_df.columns:
                    col_str = str(col)
                    for keyword in required_keywords:
                        if keyword in col_str:
                            filtered_columns.append(col)
                            break

                if filtered_columns:
                    # 创建只包含关键字列的数据
                    filtered_backup_df = backup_df[filtered_columns]

                    # 写入标题
                    for col_idx, header in enumerate(filtered_columns, 1):
                        header_cell = ws_backup.cell(row=1, column=col_idx, value=header)
                        # 设置标题格式
                        header_cell.font = Font(bold=True, color="FFFFFF")
                        header_cell.alignment = Alignment(horizontal="center", vertical="center")
                        header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

                    # 写入数据
                    for row_idx, (_, row) in enumerate(filtered_backup_df.iterrows(), 2):
                        for col_idx, value in enumerate(row, 1):
                            cell = ws_backup.cell(row=row_idx, column=col_idx, value=value)
                            # 设置数据居中对齐
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    self.logger.info(f"储备志愿者表添加完成: {len(filtered_backup_df)} 行, {len(filtered_columns)} 列")
                    self.logger.info(f"保留的字段: {', '.join(filtered_columns)}")
                else:
                    self.logger.warning("未找到任何匹配的关键字段，跳过储备志愿者表")

            # 添加颜色对照表
            ws_color = wb.create_sheet(title='颜色对照表')

            # 写入标题
            color_headers = list(color_table_df.columns)
            for col_idx, header in enumerate(color_headers, 1):
                header_cell = ws_color.cell(row=1, column=col_idx, value=header)
                # 设置标题格式
                header_cell.font = Font(bold=True, color="FFFFFF")
                header_cell.alignment = Alignment(horizontal="center", vertical="center")
                header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # 写入数据并为颜色代码列填充背景色
            for row_idx, (_, row) in enumerate(color_table_df.iterrows(), 2):
                for col_idx, (col_name, value) in enumerate(row.items(), 1):
                    cell = ws_color.cell(row=row_idx, column=col_idx, value=value)

                    # 如果是颜色代码列（第3列），填充背景色
                    if col_idx == 3 and value and str(value).strip():  # C列是颜色代码列
                        color_code = str(value).upper().lstrip('#')
                        if len(color_code) == 6:  # 确保是有效的6位十六进制颜色代码
                            try:
                                # 为颜色代码单元格填充相应的背景色
                                # 使用最简洁的颜色设置，避免任何额外的样式干扰
                                cell.fill = PatternFill(
                                    start_color=color_code,
                                    end_color=color_code,
                                    fill_type="solid"
                                )

                                # 设置字体颜色（白色背景用黑色字体，其他用白色字体）
                                if color_code.upper() == 'FFFFFF':
                                    cell.font = Font(color="000000")
                                else:
                                    cell.font = Font(color="FFFFFF")

                                # 居中对齐
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            except Exception as e:
                                self.logger.warning(f"设置颜色代码 {color_code} 的背景色失败: {str(e)}")

            # 设置列宽
            column_widths = {
                'A': 15,  # 类型
                'B': 30,  # 名称
                'C': 15,  # 颜色代码
                'D': 15,  # 颜色说明
            }
            for col, width in column_widths.items():
                ws_color.column_dimensions[col].width = width

            # 冻结首行
            ws_color.freeze_panes = 'A2'

            self.logger.info(f"颜色对照表添加完成: {len(color_table_df)} 行")

            # 保存修改后的大总表
            wb.save(integrated_file)
            wb.close()
            self.logger.info(f"大总表文件更新完成，新增sheet已保存")

        except Exception as e:
            self.logger.error(f"添加额外sheet失败: {str(e)}")

    def _create_integrated_workbook(self, integrated_file: str):
        """创建包含所有sheet的整合工作簿，重点保留总表格式"""
        try:
            self.logger.info("开始创建整合工作簿")

            # 1. 创建新的工作簿
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认sheet

            # 2. 首先复制总表，保留所有格式
            master_file = os.path.join(self.output_dir, CONFIG.get('files.master_schedule'))
            if os.path.exists(master_file):
                self._copy_master_sheet_with_format(wb, master_file, '总表')
                self.logger.info("总表复制完成")
            else:
                self.logger.warning(f"总表文件不存在: {master_file}")

            # 3. 添加小组信息表
            group_info_file = os.path.join(self.scheduling_prep_dir, CONFIG.get('files.group_info'))
            if os.path.exists(group_info_file):
                group_info_df = self.handler.read_excel(group_info_file)
                ws_group = wb.create_sheet(title='小组信息表')

                # 写入标题
                headers = list(group_info_df.columns)
                for col_idx, header in enumerate(headers, 1):
                    ws_group.cell(row=1, column=col_idx, value=header)

                # 写入数据
                for row_idx, (_, row) in enumerate(group_info_df.iterrows(), 2):
                    for col_idx, (_, value) in enumerate(row.items(), 1):
                        ws_group.cell(row=row_idx, column=col_idx, value=value)

                self.logger.info(f"小组信息表添加完成: {len(group_info_df)} 行")

            # 4. 添加储备志愿者表
            backup_file = os.path.join(self.scheduling_prep_dir, CONFIG.get('files.backup_volunteers'))
            if os.path.exists(backup_file):
                backup_df = self.handler.read_excel(backup_file)
                ws_backup = wb.create_sheet(title='储备志愿者表')

                # 写入标题
                headers = list(backup_df.columns)
                for col_idx, header in enumerate(headers, 1):
                    ws_backup.cell(row=1, column=col_idx, value=header)

                # 写入数据
                for row_idx, (_, row) in enumerate(backup_df.iterrows(), 2):
                    for col_idx, (_, value) in enumerate(row.items(), 1):
                        ws_backup.cell(row=row_idx, column=col_idx, value=value)

                self.logger.info(f"储备志愿者表添加完成: {len(backup_df)} 行")

            # 5. 添加颜色对照表
            ws_color = wb.create_sheet(title='颜色对照表')

            # 写入标题
            color_headers = list(color_table_df.columns)
            for col_idx, header in enumerate(color_headers, 1):
                ws_color.cell(row=1, column=col_idx, value=header)

            # 写入数据
            for row_idx, (_, row) in enumerate(self.color_table_df.iterrows(), 2):
                for col_idx, (_, value) in enumerate(row.items(), 1):
                    ws_color.cell(row=row_idx, column=col_idx, value=value)

            self.logger.info(f"颜色对照表添加完成: {len(self.color_table_df)} 行")

            # 保存工作簿
            wb.save(integrated_file)
            wb.close()
            self.logger.info(f"整合工作簿保存完成: {integrated_file}")

            return wb

        except Exception as e:
            self.logger.error(f"创建整合工作簿失败: {str(e)}")
            return None

    def _format_color_table(self, ws):
        """格式化颜色对照表，为颜色代码列填充背景色"""
        try:
            self.logger.info("开始格式化颜色对照表")

            # 设置列宽
            column_widths = {
                'A': 15,  # 类型
                'B': 30,  # 名称（加宽以容纳团体名称）
                'C': 12,  # 颜色代码
                'D': 15,  # 颜色说明
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # 处理每一行，为颜色代码列填充背景色
            for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
                # 跳过标题行
                if row_idx == 1:
                    continue

                # 获取各列的值
                type_cell = row[0]  # A列 - 类型
                name_cell = row[1]  # B列 - 名称
                color_code_cell = row[2]  # C列 - 颜色代码

                # 如果是分隔行（空白行），跳过
                if not type_cell.value or not name_cell.value:
                    continue

                # 处理颜色代码列的背景色填充
                if color_code_cell.value and color_code_cell.value != '':
                    color_code = str(color_code_cell.value).upper().lstrip('#')
                    if len(color_code) == 6:
                        # 为颜色代码列填充实际颜色背景
                        color_code_cell.fill = PatternFill(
                            start_color=color_code,
                            end_color=color_code,
                            fill_type="solid"
                        )

                        # 同时为白色背景的颜色代码设置黑色字体以提高可读性
                        if color_code.upper() in ['FFFFFF', 'FFFFFFF', 'FFFFFFFF']:  # 白色背景
                            color_code_cell.font = Font(color="000000", bold=True)
                        else:
                            color_code_cell.font = Font(bold=True)

                        # 设置居中对齐
                        color_code_cell.alignment = Alignment(horizontal="center", vertical="center")

            self.logger.info("颜色对照表格式化完成")

        except Exception as e:
            self.logger.error(f"格式化颜色对照表失败: {str(e)}")

    def _get_thin_border(self):
        """获取细边框样式"""
        from openpyxl.styles import Border, Side
        thin = Side(border_style="thin", color="000000")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _make_light_color(self, hex_color: str) -> str:
        """将颜色变淡，用于背景填充"""
        if len(hex_color) != 6:
            return hex_color

        try:
            # 转换为RGB
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)

            # 与白色混合，使颜色变淡
            light_r = int(r * 0.8 + 255 * 0.2)
            light_g = int(g * 0.8 + 255 * 0.2)
            light_b = int(b * 0.8 + 255 * 0.2)

            return f"{light_r:02X}{light_g:02X}{light_b:02X}"

        except:
            return hex_color

    def _copy_master_sheet_with_format(self, target_wb, source_file_path: str, target_sheet_name: str):
        """完整复制总表sheet，保留所有格式和颜色"""
        try:
            if not os.path.exists(source_file_path):
                self.logger.warning(f"原总表文件不存在: {source_file_path}")
                return

            # 加载源工作簿
            source_wb = load_workbook(source_file_path)

            if len(source_wb.sheetnames) == 0:
                self.logger.warning("原总表文件没有工作表")
                return

            # 使用第一个工作表（通常叫Sheet1）
            source_ws = source_wb.active

            # 创建目标工作表
            target_ws = target_wb.create_sheet(title=target_sheet_name)

            self.logger.info(f"开始复制总表格式，源工作表: {source_ws.title}")

            # 复制所有单元格的值和格式
            for row in source_ws.iter_rows():
                for cell in row:
                    target_cell = target_ws.cell(row=cell.row, column=cell.column)

                    # 复制值
                    target_cell.value = cell.value

                    # 复制所有样式属性
                    if cell.has_style:
                        # 复制字体
                        if cell.font:
                            target_cell.font = cell.font

                        # 复制边框
                        if cell.border:
                            target_cell.border = cell.border

                        # 复制填充（背景色）
                        if cell.fill:
                            target_cell.fill = cell.fill

                        # 复制对齐方式
                        if cell.alignment:
                            target_cell.alignment = cell.alignment

                        # 复制数字格式
                        if cell.number_format:
                            target_cell.number_format = cell.number_format

                        # 复制保护属性
                        if cell.protection:
                            target_cell.protection = cell.protection

                        # 复制文本方向
                        if cell.text_rotation:
                            target_cell.text_rotation = cell.text_rotation

            # 复制列宽
            for col_letter, dimension in source_ws.column_dimensions.items():
                target_ws.column_dimensions[col_letter].width = dimension.width

            # 复制行高
            for row_num, dimension in source_ws.row_dimensions.items():
                target_ws.row_dimensions[row_num].height = dimension.height

            # 复制合并单元格
            if source_ws.merged_cells:
                for merged_range in source_ws.merged_cells.ranges:
                    target_ws.merge_cells(str(merged_range))

            # 复制冻结窗格
            if source_ws.freeze_panes:
                target_ws.freeze_panes = source_ws.freeze_panes

            source_wb.close()
            self.logger.info(f"成功复制总表到 {target_sheet_name}，包含所有格式和颜色")

        except Exception as e:
            self.logger.error(f"复制总表格式失败: {str(e)}")
            # 如果复制失败，至少复制数据
            try:
                master_df = self.handler.read_excel(source_file_path)
                with pd.ExcelWriter(os.path.join(self.output_dir, 'temp_master.xlsx'), engine='openpyxl') as writer:
                    master_df.to_excel(writer, sheet_name=target_sheet_name, index=False)

                temp_wb = load_workbook(os.path.join(self.output_dir, 'temp_master.xlsx'))
                temp_ws = temp_wb.active
                target_ws = target_wb.create_sheet(title=target_sheet_name + "_backup")

                for row in temp_ws.iter_rows():
                    for cell in row:
                        target_ws.cell(row=cell.row, column=cell.column, value=cell.value)

                temp_wb.close()
                os.remove(os.path.join(self.output_dir, 'temp_master.xlsx'))
                self.logger.info("使用备份数据复制方法")

            except Exception as backup_e:
                self.logger.error(f"备份数据复制也失败: {str(backup_e)}")

    def _format_integrated_file(self, file_path: str):
        """美化整合文件格式"""
        try:
            # 加载工作簿
            wb = load_workbook(file_path)

            # 只为新添加的sheet设置格式，跳过总表以保留其原有格式
            new_sheets = ['小组信息表', '储备志愿者表', '颜色对照表']
            for sheet_name in wb.sheetnames:
                if sheet_name not in new_sheets:
                    continue  # 跳过总表，保留其原有格式

                ws = wb[sheet_name]

                # 设置标题行格式
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(horizontal="center", vertical="center")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

                # 设置标题行样式
                for cell in ws[1]:  # 第一行
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill

                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
                    ws.column_dimensions[column_letter].width = adjusted_width

                # 设置数据对齐
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # 冻结首行
                ws.freeze_panes = 'A2'

            # 保存工作簿
            wb.save(file_path)
            self.logger.info("整合文件格式美化完成")

        except Exception as e:
            self.logger.warning(f"美化整合文件失败: {str(e)}")

    def _calculate_finalization_statistics(self, split_files: Dict[int, str], integrated_file: str) -> Dict[str, Any]:
        """计算最终处理统计信息"""
        self.logger.info("计算最终处理统计信息")

        statistics = {
            'total_groups': len(split_files),
            'split_files_count': len(split_files),
            'groups_with_files': list(split_files.keys()),
            'integrated_file_generated': integrated_file is not None,
            'group_details': {}
        }

        # 统计每个小组的详细信息
        for group_number, file_path in split_files.items():
            try:
                df = self.handler.read_excel(file_path)
                statistics['group_details'][group_number] = {
                    'member_count': len(df),
                    'file_path': file_path,
                    'file_size': os.path.getsize(file_path)
                }
            except Exception as e:
                self.logger.warning(f"统计小组 {group_number} 信息失败: {str(e)}")
                statistics['group_details'][group_number] = {
                    'member_count': 0,
                    'file_path': file_path,
                    'error': str(e)
                }

        # 计算总体统计
        total_members = sum(info['member_count'] for info in statistics['group_details'].values())
        statistics['total_members_in_groups'] = total_members

        if integrated_file and os.path.exists(integrated_file):
            statistics['integrated_file_size'] = os.path.getsize(integrated_file)

        self.logger.info(f"最终处理统计：{len(split_files)} 个小组文件，{total_members} 名志愿者")
        return statistics

    def validate_split_files(self, split_files: Dict[int, str], master_df: pd.DataFrame) -> bool:
        """验证拆分文件的完整性"""
        self.logger.info("验证拆分文件完整性")

        try:
            # 检查总人数
            total_in_splits = 0
            for group_number, file_path in split_files.items():
                df = self.handler.read_excel(file_path)
                total_in_splits += len(df)

            total_in_master = len(master_df)

            if total_in_splits != total_in_master:
                self.logger.error(f"人数不匹配：拆分文件总计 {total_in_splits} 人，总表 {total_in_master} 人")
                return False

            # 检查小组号连续性
            expected_groups = set(master_df['小组号'].unique())
            actual_groups = set(split_files.keys())

            if expected_groups != actual_groups:
                self.logger.error(f"小组号不匹配：期望 {expected_groups}，实际 {actual_groups}")
                return False

            # 检查文件格式
            for group_number, file_path in split_files.items():
                try:
                    df = self.handler.read_excel(file_path)

                    # 检查必要列
                    required_columns = ['小组号', '学号', '姓名']
                    missing_columns = [col for col in required_columns if col not in df.columns]

                    if missing_columns:
                        self.logger.error(f"小组 {group_number} 文件缺少必要列: {missing_columns}")
                        return False

                    # 检查敏感信息是否已移除
                    sensitive_columns = ['证件类型', '证件号']
                    found_sensitive = [col for col in sensitive_columns if col in df.columns]

                    if found_sensitive:
                        self.logger.warning(f"小组 {group_number} 文件仍包含敏感信息: {found_sensitive}")

                except Exception as e:
                    self.logger.error(f"验证小组 {group_number} 文件失败: {str(e)}")
                    return False

            self.logger.info("拆分文件验证通过")
            return True

        except Exception as e:
            self.logger.error(f"验证拆分文件失败: {str(e)}")
            return False

    def cleanup_temp_files(self):
        """清理临时文件"""
        self.logger.info("清理临时文件")
        # 这里可以添加清理逻辑，比如删除临时文件等
        pass


def main():
    """命令行入口函数"""
    import argparse

    parser = argparse.ArgumentParser(description='总表拆分和整合程序')
    parser.add_argument('--master-file', help='总表文件路径')
    parser.add_argument('--output-dir', help='输出目录路径')
    parser.add_argument('--split-only', action='store_true', help='仅执行拆分，不生成整合文件')
    parser.add_argument('--integrate-only', action='store_true', help='仅生成整合文件，不执行拆分')

    args = parser.parse_args()

    logger = get_logger(__file__)
    logger.info("开始执行总表拆分和整合程序")

    try:
        finalizer = Finalizer()

        # 如果指定了自定义路径，更新配置
        if args.output_dir:
            finalizer.output_dir = args.output_dir
            finalizer.groups_output_dir = os.path.join(args.output_dir, '各小组名单')

        # 执行最终处理
        if args.split_only:
            # 仅执行拆分
            master_df = finalizer._read_master_schedule()
            split_files = finalizer._split_master_schedule(master_df)

            stats = finalizer._calculate_finalization_statistics(split_files, None)
            print(f"\n📊 拆分结果:")
            print(f"  拆分文件数: {stats['split_files_count']}")
            print(f"  总志愿者数: {stats['total_members_in_groups']} 人")

        elif args.integrate_only:
            # 仅生成整合文件
            integrated_file = finalizer._generate_integrated_schedule()
            print(f"\n📄 整合文件已生成: {integrated_file}")

        else:
            # 执行完整流程
            results = finalizer.run_finalization()

            # 输出结果摘要
            if not results['errors']:
                stats = results['statistics']
                print(f"\n📊 最终处理结果:")
                print(f"  拆分小组数: {stats['total_groups']} 个")
                print(f"  小组文件数: {stats['split_files_count']} 个")
                print(f"  总志愿者数: {stats['total_members_in_groups']} 人")
                print(f"  整合文件: {'已生成' if stats['integrated_file_generated'] else '未生成'}")

                if stats['integrated_file_generated']:
                    file_size_mb = stats.get('integrated_file_size', 0) / (1024 * 1024)
                    print(f"  整合文件大小: {file_size_mb:.2f} MB")

                print(f"\n📁 输出目录:")
                print(f"  小组文件: {finalizer.groups_output_dir}")
                print(f"  整合文件: {results['integrated_file']}")

                # 显示小组详情
                print(f"\n📋 小组详情:")
                for group_number, details in stats['group_details'].items():
                    print(f"  小组 {group_number}: {details['member_count']} 人")

            else:
                print(f"\n❌ 最终处理失败:")
                for error in results['errors']:
                    print(f"  - {error}")

    except Exception as e:
        logger.error(f"程序执行失败: {str(e)}")
        sys.exit(1)


if __name__ == '__main__':
    main()