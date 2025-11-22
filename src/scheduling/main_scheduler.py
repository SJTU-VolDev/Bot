"""
排表主程序 - 核心调度引擎
负责将所有志愿者分配到各个小组中，生成最终的排班总表

采用"内存对象模型"策略，先在内存中完成所有分配逻辑，最后统一写入Excel
支持"先整体后个体，先委派后自由"的分配原则
"""

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Dict, List, Set, Optional, Tuple, Any
import logging
from dataclasses import dataclass
import pandas as pd

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent
sys.path.append(str(project_root))

from src.utils.logger_factory import get_logger
from src.utils._excel_handler import ExcelHandler
from src.scheduling.data_models import (
    Volunteer, Group, Position, BindingSet, SchedulingMetadata,
    VolunteerType, SpecialRole, DirectAssignment
)
from config.loader import CONFIG, get_file_path


@dataclass
class AssignmentResult:
    """分配结果"""
    success: bool
    message: str
    assigned_count: int = 0
    failed_assignments: List[str] = None

    def __post_init__(self):
        if self.failed_assignments is None:
            self.failed_assignments = []


class VolunteerScheduler:
    """志愿者排表调度器"""

    def __init__(self):
        """初始化调度器"""
        self.logger = get_logger(__file__)
        self.excel_handler = ExcelHandler()

        # 核心数据结构
        self.groups: Dict[int, Group] = {}
        self.volunteers: Dict[str, Volunteer] = {}  # 学号 -> 志愿者
        self.binding_sets: List[BindingSet] = []
        self.direct_assignments: Dict[str, int] = {}  # 学号 -> 小组号
        self.metadata: SchedulingMetadata = SchedulingMetadata()

        # 分配状态追踪
        self.placed_student_ids: Set[str] = set()
        self.group_colors: Dict[str, str] = {}  # 团体名称 -> 颜色

        # 加载配置
        self._load_colors()

    def _load_colors(self):
        """加载颜色配置"""
        colors = CONFIG.get('colors', {})
        self.leader_color = colors.get('leader', 'FFFF00')
        self.lightning_color = colors.get('lightning', '00FF00')
        self.photography_color = colors.get('photography', 'E6E6FA')
        self.couple_color = colors.get('couple', 'FFB6C1')
        self.internal_color = colors.get('internal', 'FFA500')
        self.family_color = colors.get('family', '87CEEB')
        self.default_color = colors.get('default', 'FFFFFF')

        # 团体颜色
        self.group_color_list = colors.get('group_colors', [
            '98FB98', 'DDA0DD', 'F0E68C', 'ADD8E6', 'F5DEB3'
        ])

    def _find_column_by_keyword(self, columns, keyword: str) -> Optional[str]:
        """根据关键字查找列名"""
        if not keyword:
            return None

        keyword = keyword.lower()
        # 处理 DataFrame 和 Index 的情况
        if hasattr(columns, 'columns'):  # 是 DataFrame
            col_iter = columns.columns
        else:  # 是 Index 或其他序列
            col_iter = columns

        for col in col_iter:
            if keyword in str(col).lower():
                return col
        return None

    def _extract_volunteer_from_row(self, row: pd.Series, volunteer_type: VolunteerType) -> Volunteer:
        """从数据行中提取志愿者信息"""
        # 基础字段映射
        field_mappings = {
            'student_id': CONFIG.get('field_mappings.student_id', '学号'),
            'name': CONFIG.get('field_mappings.name', '姓名'),
            'name_pinyin': CONFIG.get('field_mappings.name_pinyin', '姓名拼音'),
            'gender': CONFIG.get('field_mappings.gender', '性别'),
            'id_type': CONFIG.get('field_mappings.id_type', '证件类型'),
            'id_number': CONFIG.get('field_mappings.id_number', '证件号'),
            'birth_date': CONFIG.get('field_mappings.birth_date', '出生日期'),
            'college': CONFIG.get('field_mappings.college', '学院'),
            'height': CONFIG.get('field_mappings.height', '身高'),
            'email': CONFIG.get('field_mappings.email', '邮箱'),
            'phone': CONFIG.get('field_mappings.phone', '手机号'),
            'qq': CONFIG.get('field_mappings.qq', 'QQ号'),
            'wechat': CONFIG.get('field_mappings.wechat', '微信号'),
            'political_status': CONFIG.get('field_mappings.political_status', '政治面貌'),
            'marathon_count': CONFIG.get('field_mappings.marathon_count', '第几次做马拉松志愿者'),
            'campus': CONFIG.get('field_mappings.campus', '校区'),
            'clothes_size': CONFIG.get('field_mappings.clothes_size', '衣服尺码')
        }

        # 提取基础信息
        kwargs = {}
        for field, keyword in field_mappings.items():
            col_name = self._find_column_by_keyword(row.index, keyword)
            if col_name and pd.notna(row[col_name]):
                kwargs[field] = str(row[col_name]).strip()

        # 处理宿舍楼栋逻辑
        dorm_minhang_col = self._find_column_by_keyword(row.index,
            CONFIG.get('field_mappings.dorm_building_minhang', '宿舍楼栋（闵行）'))
        dorm_non_minhang_col = self._find_column_by_keyword(row.index,
            CONFIG.get('field_mappings.dorm_building_non_minhang', '宿舍楼栋（非闵行）'))

        if dorm_minhang_col and pd.notna(row[dorm_minhang_col]) and str(row[dorm_minhang_col]).strip():
            kwargs['dorm_building'] = str(row[dorm_minhang_col]).strip()
        elif dorm_non_minhang_col and pd.notna(row[dorm_non_minhang_col]):
            kwargs['dorm_building'] = str(row[dorm_non_minhang_col]).strip()

        # 必须有学号和姓名
        if 'student_id' not in kwargs or 'name' not in kwargs:
            raise ValueError("学号和姓名是必需的")

        # 创建志愿者对象
        volunteer = Volunteer(volunteer_type=volunteer_type, **kwargs)

        return volunteer

    def _calculate_background_color(self, volunteer: Volunteer) -> str:
        """计算志愿者背景颜色（身份 > 属性原则）"""
        # 优先级：组长 > 小闪电 > 摄影 > 情侣
        if volunteer.has_special_role(SpecialRole.LEADER):
            return self.leader_color
        elif volunteer.has_special_role(SpecialRole.LIGHTNING):
            return self.lightning_color
        elif volunteer.has_special_role(SpecialRole.PHOTOGRAPHY):
            return self.photography_color
        elif volunteer.has_special_role(SpecialRole.COUPLE):
            return self.couple_color

        # 属性颜色
        if volunteer.volunteer_type == VolunteerType.INTERNAL:
            return self.internal_color
        elif volunteer.volunteer_type == VolunteerType.FAMILY:
            return self.family_color
        elif volunteer.volunteer_type == VolunteerType.GROUP:
            return self.group_colors.get(volunteer.group_name, self.group_color_list[0])

        # 默认无背景色
        return self.default_color

    def load_data(self) -> bool:
        """加载所有输入数据"""
        try:
            self.logger.info("开始加载数据...")

            # 1. 加载元数据
            if not self._load_metadata():
                return False

            # 2. 加载志愿者数据（先加载志愿者，这样小组信息加载时就能匹配组长）
            if not self._load_volunteers():
                return False

            # 3. 加载小组信息（现在可以在加载时分配组长）
            if not self._load_groups():
                return False

            # 4. 加载绑定集合
            if not self._load_binding_sets():
                return False

            # 5. 加载直接委派名单
            if not self._load_direct_assignments():
                return False

            # 6. 加载情侣信息
            if not self._load_couple_info():
                return False

            self.logger.info("数据加载完成")
            return True

        except Exception as e:
            self.logger.error(f"数据加载失败: {str(e)}")
            return False

    def _load_metadata(self) -> bool:
        """加载元数据"""
        try:
            metadata_path = get_file_path('metadata')
            if not os.path.exists(metadata_path):
                self.logger.warning(f"元数据文件不存在: {metadata_path}")
                return True

            with open(metadata_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # 更新元数据对象
            for key, value in data.items():
                if hasattr(self.metadata, key):
                    setattr(self.metadata, key, value)

            self.logger.info(f"元数据加载成功: {data.get('total_required_volunteers', 'N/A')} 个岗位需求")
            return True

        except Exception as e:
            self.logger.error(f"元数据加载失败: {str(e)}")
            return False

    def _load_groups(self) -> bool:
        """加载小组信息"""
        try:
            group_info_path = get_file_path('group_info')
            self.logger.info(f"尝试读取小组信息文件: {group_info_path}")

            if not os.path.exists(group_info_path):
                self.logger.error(f"小组信息文件不存在: {group_info_path}")
                return False

            df = self.excel_handler.read_excel(group_info_path)
            self.logger.info(f"小组信息表列名: {df.columns.tolist()}")
            self.logger.info(f"小组信息表行数: {len(df)}")

            # 查找必要的列
            group_id_col = self._find_column_by_keyword(df, CONFIG.get('field_mappings.group_number', '小组号'))
            position_name_col = self._find_column_by_keyword(df, CONFIG.get('field_mappings.position_name', '岗位名称'))
            position_desc_col = self._find_column_by_keyword(df, CONFIG.get('field_mappings.position_description', '岗位简介'))
            leader_name_col = self._find_column_by_keyword(df, CONFIG.get('field_mappings.group_leader', '组长姓名'))
            leader_id_col = self._find_column_by_keyword(df, '组长学号')
            required_count_col = self._find_column_by_keyword(df, CONFIG.get('field_mappings.required_count', '小组人数'))

            self.logger.info(f"列查找结果: group_id_col={group_id_col}, position_name_col={position_name_col}, leader_name_col={leader_name_col}, leader_id_col={leader_id_col}, required_count_col={required_count_col}")

            if not all([group_id_col, position_name_col, required_count_col]):
                missing_cols = [col for col, found in [
                    ('小组号', group_id_col), ('岗位名称', position_name_col), ('小组人数', required_count_col)
                ] if not found]
                self.logger.error(f"无法找到必要的列: {missing_cols}")
                return False

            groups_assigned_leaders = 0
            for _, row in df.iterrows():
                group_id = int(row[group_id_col])
                position_name = str(row[position_name_col]).strip()
                position_desc = str(row[position_desc_col]).strip() if pd.notna(row[position_desc_col]) and position_desc_col else ""
                leader_name = str(row[leader_name_col]).strip() if pd.notna(row[leader_name_col]) and leader_name_col else ""
                leader_student_id = str(row[leader_id_col]).strip() if pd.notna(row[leader_id_col]) and leader_id_col else ""
                required_count = int(row[required_count_col])

                # 创建岗位对象
                position = Position(
                    name=position_name,
                    description=position_desc,
                    required_count=required_count
                )

                # 创建小组对象
                group = Group(
                    group_id=group_id,
                    position=position,
                    required_count=required_count
                )

                # 尝试立即分配组长
                leader = None
                if leader_student_id and leader_student_id in self.volunteers:
                    # 优先通过学号匹配
                    leader = self.volunteers[leader_student_id]
                    self.logger.info(f"小组 {group_id} 通过学号匹配找到组长: {leader.name} ({leader_student_id})")
                elif leader_name:
                    # 通过姓名匹配
                    for volunteer in self.volunteers.values():
                        if volunteer.name == leader_name:
                            leader = volunteer
                            self.logger.info(f"小组 {group_id} 通过姓名匹配找到组长: {leader.name} ({volunteer.student_id})")
                            break

                if leader:
                    # 设置组长信息并立即分配
                    leader.add_special_role(SpecialRole.LEADER)
                    leader.is_leader = True
                    group.leader = leader
                    group.add_member(leader)
                    self.placed_student_ids.add(leader.student_id)
                    groups_assigned_leaders += 1
                    self.logger.info(f"小组 {group_id} 组长分配成功: {leader.name} ({leader.student_id})")
                else:
                    if leader_name or leader_student_id:  # 只有在有组长信息时才警告
                        self.logger.warning(f"小组 {group_id} 未找到匹配的组长: 姓名='{leader_name}', 学号='{leader_student_id}'")

                self.groups[group_id] = group

            self.logger.info(f"小组信息加载成功: {len(self.groups)} 个小组，其中 {groups_assigned_leaders} 个小组成功分配组长")
            return True

        except Exception as e:
            self.logger.error(f"小组信息加载失败: {str(e)}")
            return False

    def _load_volunteers(self) -> bool:
        """加载志愿者数据"""
        try:
            # 1. 加载正式普通志愿者
            if not self._load_formal_normal_volunteers():
                return False

            # 2. 加载内部志愿者
            if not self._load_internal_volunteers():
                return False

            # 3. 加载家属志愿者
            if not self._load_family_volunteers():
                return False

            # 4. 加载团体志愿者
            if not self._load_group_volunteers():
                return False

            # 5. 加载面试成绩信息
            if not self._load_interview_scores():
                return False

            total_volunteers = len(self.volunteers)
            self.logger.info(f"志愿者数据加载成功: {total_volunteers} 个志愿者")

            # 按类型统计
            type_stats = {}
            for volunteer in self.volunteers.values():
                type_name = volunteer.volunteer_type.value
                type_stats[type_name] = type_stats.get(type_name, 0) + 1

            for vtype, count in type_stats.items():
                self.logger.info(f"  {vtype}: {count} 个")

            return True

        except Exception as e:
            self.logger.error(f"志愿者数据加载失败: {str(e)}")
            return False

    def _load_formal_normal_volunteers(self) -> bool:
        """加载正式普通志愿者"""
        try:
            formal_path = get_file_path('formal_normal_volunteers')
            self.logger.info(f"尝试读取正式普通志愿者文件: {formal_path}")

            if not os.path.exists(formal_path):
                self.logger.warning("正式普通志愿者表不存在，跳过")
                return True

            df = self.excel_handler.read_excel(formal_path)
            self.logger.info(f"正式普通志愿者表列名: {df.columns.tolist()}")
            self.logger.info(f"正式普通志愿者表行数: {len(df)}")

            for idx, row in df.iterrows():
                try:
                    volunteer = self._extract_volunteer_from_row(row, VolunteerType.NORMAL)
                    if volunteer.student_id not in self.volunteers:
                        self.volunteers[volunteer.student_id] = volunteer
                    else:
                        self.logger.warning(f"学号重复 {volunteer.student_id}，跳过")
                except Exception as e:
                    self.logger.warning(f"处理第 {idx+1} 行数据时出错: {str(e)}")
                    continue

            self.logger.info(f"正式普通志愿者加载完成: {len(df)} 个")
            return True

        except Exception as e:
            self.logger.error(f"正式普通志愿者加载失败: {str(e)}")
            return False

    def _load_internal_volunteers(self) -> bool:
        """加载内部志愿者"""
        try:
            internal_path = get_file_path('internal_volunteers')
            df = self.excel_handler.read_excel(internal_path)

            for _, row in df.iterrows():
                volunteer = self._extract_volunteer_from_row(row, VolunteerType.INTERNAL)

                # 检查是否是组长
                leader_col = self._find_column_by_keyword(row,
                    CONFIG.get('field_mappings.leader_role', '小组长或者区长'))
                if leader_col and pd.notna(row[leader_col]):
                    leader_role = str(row[leader_col]).strip()
                    if leader_role in ['小组长', '区长']:
                        volunteer.add_special_role(SpecialRole.LEADER)
                        volunteer.is_leader = True

                self.volunteers[volunteer.student_id] = volunteer

            self.logger.info(f"内部志愿者加载完成: {len(df)} 个")
            return True

        except Exception as e:
            self.logger.error(f"内部志愿者加载失败: {str(e)}")
            return False

    def _load_family_volunteers(self) -> bool:
        """加载家属志愿者"""
        try:
            family_path = get_file_path('family_volunteers')
            df = self.excel_handler.read_excel(family_path)

            for _, row in df.iterrows():
                volunteer = self._extract_volunteer_from_row(row, VolunteerType.FAMILY)

                # 家属相关信息
                related_col = self._find_column_by_keyword(row,
                    CONFIG.get('field_mappings.family_of', '您是谁的家属'))
                hope_same_col = self._find_column_by_keyword(row,
                    CONFIG.get('field_mappings.hope_same_group', '是否希望与他/她同组'))

                if related_col and pd.notna(row[related_col]):
                    volunteer.related_internal_name = str(row[related_col]).strip()
                if hope_same_col and pd.notna(row[hope_same_col]):
                    volunteer.hope_same_group = str(row[hope_same_col]).strip() == '是'

                self.volunteers[volunteer.student_id] = volunteer

            self.logger.info(f"家属志愿者加载完成: {len(df)} 个")
            return True

        except Exception as e:
            self.logger.error(f"家属志愿者加载失败: {str(e)}")
            return False

    def _load_group_volunteers(self) -> bool:
        """加载团体志愿者"""
        try:
            groups_dir = CONFIG.get('paths.groups_dir')
            if not os.path.exists(groups_dir):
                self.logger.info("团体志愿者目录不存在，跳过")
                return True

            group_files = [f for f in os.listdir(groups_dir) if f.endswith(('.xlsx', '.xls'))]
            color_index = 0

            for group_file in group_files:
                group_name = os.path.splitext(group_file)[0]
                file_path = os.path.join(groups_dir, group_file)

                # 分配颜色
                if group_name not in self.group_colors:
                    self.group_colors[group_name] = self.group_color_list[color_index % len(self.group_color_list)]
                    color_index += 1

                df = self.excel_handler.read_excel(file_path)
                for _, row in df.iterrows():
                    volunteer = self._extract_volunteer_from_row(row, VolunteerType.GROUP)
                    volunteer.group_name = group_name
                    self.volunteers[volunteer.student_id] = volunteer

            self.logger.info(f"团体志愿者加载完成: {len(group_files)} 个团体")

            # 更新metadata中的团体颜色信息
            self.metadata.group_color_mapping = self.group_colors.copy()
            self.logger.info(f"团体颜色映射已更新: {len(self.group_colors)} 个团体")

            return True

        except Exception as e:
            self.logger.error(f"团体志愿者加载失败: {str(e)}")
            return False

    def _load_interview_scores(self) -> bool:
        """加载面试成绩信息"""
        try:
            scores_path = get_file_path('unified_interview_scores')
            self.logger.info(f"尝试读取统一面试打分表文件: {scores_path}")

            if not os.path.exists(scores_path):
                self.logger.warning("统一面试打分表不存在，跳过")
                return True

            df = self.excel_handler.read_excel(scores_path)
            self.logger.info(f"统一面试打分表列名: {df.columns.tolist()}")
            self.logger.info(f"统一面试打分表行数: {len(df)}")

            # 查找相关列
            name_col = self._find_column_by_keyword(df, '姓名')
            student_id_col = self._find_column_by_keyword(df, '学号')
            score_col = self._find_column_by_keyword(df,
                CONFIG.get('field_mappings.normalized_score', '归一化成绩'))
            lightning_col = self._find_column_by_keyword(df,
                CONFIG.get('field_mappings.lightning_score', '小闪电得分'))
            photography_col = self._find_column_by_keyword(df,
                CONFIG.get('field_mappings.photography_score', '摄影得分'))

            self.logger.info(f"面试成绩列查找结果: name_col={name_col}, student_id_col={student_id_col}, score_col={score_col}, lightning_col={lightning_col}, photography_col={photography_col}")

            updated_count = 0
            lightning_count = 0
            photography_count = 0

            for _, row in df.iterrows():
                # 通过姓名或学号匹配志愿者
                student_id = str(row[student_id_col]).strip() if student_id_col else ""
                name = str(row[name_col]).strip() if name_col else ""

                volunteer = None
                if student_id and student_id in self.volunteers:
                    volunteer = self.volunteers[student_id]
                elif name:
                    # 如果学号匹配失败，尝试姓名匹配
                    for v in self.volunteers.values():
                        if v.name == name:
                            volunteer = v
                            break

                if not volunteer:
                    continue

                # 设置成绩
                if score_col and pd.notna(row[score_col]):
                    try:
                        volunteer.normalized_score = float(row[score_col])
                        updated_count += 1
                    except (ValueError, TypeError):
                        pass

                if lightning_col and pd.notna(row[lightning_col]):
                    try:
                        score = float(row[lightning_col])
                        if score > 0:
                            volunteer.lightning_score = score
                            lightning_count += 1
                    except (ValueError, TypeError):
                        pass

                if photography_col and pd.notna(row[photography_col]):
                    try:
                        score = float(row[photography_col])
                        if score > 0:
                            volunteer.photography_score = score
                            photography_count += 1
                    except (ValueError, TypeError):
                        pass

            self.logger.info(f"面试成绩信息加载完成: 更新{updated_count}人, 小闪电{lightning_count}人, 摄影{photography_count}人")
            return True

        except Exception as e:
            self.logger.error(f"面试成绩信息加载失败: {str(e)}")
            return False

    def _load_binding_sets(self) -> bool:
        """加载绑定集合"""
        try:
            binding_path = get_file_path('binding_sets')
            self.logger.info(f"尝试读取绑定集合文件: {binding_path}")

            if not os.path.exists(binding_path):
                self.logger.warning("绑定集合表不存在，跳过")
                return True

            df = self.excel_handler.read_excel(binding_path)
            self.logger.info(f"绑定集合表列名: {df.columns.tolist()}")
            self.logger.info(f"绑定集合表行数: {len(df)}")

            # 查找必要的列
            binding_id_col = self._find_column_by_keyword(df, CONFIG.get('field_mappings.binding_id', '绑定集合ID'))
            member_id_col = self._find_column_by_keyword(df, CONFIG.get('field_mappings.member_student_id', '成员学号'))
            target_group_col = self._find_column_by_keyword(df, CONFIG.get('field_mappings.target_group', '目标小组'))

            self.logger.info(f"绑定集合列查找结果: binding_id_col={binding_id_col}, member_id_col={member_id_col}, target_group_col={target_group_col}")

            if not all([binding_id_col, member_id_col]):
                missing_cols = [col for col, found in [
                    ('绑定集合ID', binding_id_col), ('成员学号', member_id_col)
                ] if not found]
                self.logger.error(f"无法找到绑定集合必要的列: {missing_cols}")
                return False

            # 按绑定集合ID分组
            binding_groups = df.groupby(binding_id_col)

            for binding_id, group_df in binding_groups:
                binding_set = BindingSet(binding_id=str(binding_id))

                # 检查是否有目标小组
                if target_group_col and pd.notna(group_df[target_group_col].iloc[0]):
                    binding_set.target_group_id = int(group_df[target_group_col].iloc[0])

                # 添加成员
                for _, row in group_df.iterrows():
                    student_id = str(row[member_id_col]).strip()
                    if student_id in self.volunteers:
                        binding_set.add_member(self.volunteers[student_id])
                    else:
                        self.logger.warning(f"绑定集合中找不到志愿者: {student_id}")

                if binding_set.members:  # 只添加非空的绑定集合
                    self.binding_sets.append(binding_set)

            self.logger.info(f"绑定集合加载完成: {len(self.binding_sets)} 个")
            return True

        except Exception as e:
            self.logger.error(f"绑定集合加载失败: {str(e)}")
            return False

    def _load_direct_assignments(self) -> bool:
        """加载直接委派名单"""
        try:
            direct_path = get_file_path('direct_assignments')
            if not os.path.exists(direct_path):
                self.logger.warning("直接委派名单不存在，跳过")
                return True

            df = self.excel_handler.read_excel(direct_path)

            student_id_col = self._find_column_by_keyword(df, CONFIG.get('field_mappings.student_id', '学号'))
            group_id_col = self._find_column_by_keyword(df, CONFIG.get('field_mappings.group_number', '小组号'))

            for _, row in df.iterrows():
                student_id = str(row[student_id_col]).strip()
                group_id = int(row[group_id_col])
                self.direct_assignments[student_id] = group_id

                # 标记志愿者
                if student_id in self.volunteers:
                    self.volunteers[student_id].is_direct_assigned = True

            self.logger.info(f"直接委派名单加载完成: {len(self.direct_assignments)} 个")
            return True

        except Exception as e:
            self.logger.error(f"直接委派名单加载失败: {str(e)}")
            return False

    def _load_couple_info(self) -> bool:
        """加载情侣信息"""
        try:
            couple_path = get_file_path('couple_volunteers')
            if not os.path.exists(couple_path):
                self.logger.warning("情侣志愿者表不存在，跳过")
                return True

            df = self.excel_handler.read_excel(couple_path)

            couple1_id_col = self._find_column_by_keyword(df,
                CONFIG.get('field_mappings.couple1_student_id', '情侣一学号'))
            couple1_name_col = self._find_column_by_keyword(df,
                CONFIG.get('field_mappings.couple1_name', '情侣一姓名'))
            couple2_id_col = self._find_column_by_keyword(df,
                CONFIG.get('field_mappings.couple2_student_id', '情侣二学号'))
            couple2_name_col = self._find_column_by_keyword(df,
                CONFIG.get('field_mappings.couple2_name', '情侣二姓名'))

            for _, row in df.iterrows():
                couple1_id = str(row[couple1_id_col]).strip()
                couple2_id = str(row[couple2_id_col]).strip()

                # 为双方添加情侣身份
                if couple1_id in self.volunteers:
                    volunteer = self.volunteers[couple1_id]
                    volunteer.add_special_role(SpecialRole.COUPLE)
                    volunteer.couple_student_id = couple2_id
                    if couple1_name_col and pd.notna(row[couple1_name_col]):
                        volunteer.couple_name = str(row[couple1_name_col]).strip()

                if couple2_id in self.volunteers:
                    volunteer = self.volunteers[couple2_id]
                    volunteer.add_special_role(SpecialRole.COUPLE)
                    volunteer.couple_student_id = couple1_id
                    if couple2_name_col and pd.notna(row[couple2_name_col]):
                        volunteer.couple_name = str(row[couple2_name_col]).strip()

            self.logger.info("情侣信息加载完成")
            return True

        except Exception as e:
            self.logger.error(f"情侣信息加载失败: {str(e)}")
            return False

    def execute_scheduling(self) -> AssignmentResult:
        """执行排表调度"""
        try:
            self.logger.info("开始执行排表调度...")

            # 阶段一：组长分配
            if not self._assign_leaders():
                return AssignmentResult(False, "组长分配失败")

            # 阶段二：绑定集合分配
            if not self._assign_binding_sets():
                return AssignmentResult(False, "绑定集合分配失败")

            # 阶段三：直接委派的落单个人
            if not self._assign_direct_individuals():
                return AssignmentResult(False, "直接委派个人分配失败")

            # 阶段四：小闪电志愿者分配（按文档要求在落单志愿者之前）
            if not self._assign_lightning_volunteers_properly():
                return AssignmentResult(False, "小闪电志愿者分配失败")

            # 阶段五：摄影志愿者分配（按文档要求在落单志愿者之前）
            if not self._assign_photography_volunteers_properly():
                return AssignmentResult(False, "摄影志愿者分配失败")

            # 阶段六：剩余内部与家属志愿者
            if not self._assign_remaining_internal_family():
                return AssignmentResult(False, "剩余内部/家属志愿者分配失败")

            # 阶段七：最终填充
            if not self._final_fill():
                return AssignmentResult(False, "最终填充失败")

            assigned_count = len(self.placed_student_ids)
            self.logger.info(f"排表调度完成: {assigned_count} 个志愿者已分配")

            return AssignmentResult(True, "排表调度成功", assigned_count)

        except Exception as e:
            self.logger.error(f"排表调度失败: {str(e)}")
            return AssignmentResult(False, f"排表调度异常: {str(e)}")

    def _assign_leaders(self) -> bool:
        """阶段一：组长分配"""
        try:
            self.logger.info("阶段一：检查组长分配情况...")

            # 统计已有组长的小组
            groups_with_leaders = 0
            groups_without_leaders = []

            for group_id, group in self.groups.items():
                if group.leader is not None:
                    groups_with_leaders += 1
                    self.logger.info(f"小组 {group_id} 已有组长: {group.leader.name} ({group.leader.student_id})")
                else:
                    groups_without_leaders.append(group_id)

            self.logger.info(f"组长分配检查完成: {groups_with_leaders} 个小组已有组长，{len(groups_without_leaders)} 个小组缺少组长")

            # 对于没有组长的小组，尝试从剩余的内部志愿者中分配
            if groups_without_leaders:
                leader_candidates = [v for v in self.volunteers.values()
                                   if v.has_special_role(SpecialRole.LEADER) and
                                   v.assigned_group_id is None]

                self.logger.info(f"找到 {len(leader_candidates)} 个可用的组长候选人")

                for group_id in groups_without_leaders:
                    if not leader_candidates:
                        self.logger.warning(f"小组 {group_id} 没有可用的组长候选人")
                        continue

                    group = self.groups[group_id]
                    # 选择第一个候选人
                    leader = leader_candidates.pop(0)

                    # 将组长分配到小组
                    group.leader = leader  # 设置小组的组长
                    group.add_member(leader)
                    self.placed_student_ids.add(leader.student_id)

                    self.logger.info(f"组长 {leader.name} ({leader.student_id}) 分配到小组 {group_id}")

            self.logger.info("组长分配完成")
            return True

        except Exception as e:
            self.logger.error(f"组长分配失败: {str(e)}")
            return False

    def _assign_binding_sets(self) -> bool:
        """阶段二：绑定集合分配"""
        try:
            self.logger.info("阶段二：开始分配绑定集合...")

            # 分离直接委派和自由分配的绑定集合
            priority_bindings = [bs for bs in self.binding_sets if bs.target_group_id is not None]
            free_bindings = [bs for bs in self.binding_sets if bs.target_group_id is None]

            # 按集合大小降序排序（大集合优先）
            free_bindings.sort(key=lambda x: x.get_size(), reverse=True)

            # 处理直接委派的绑定集合
            for binding in priority_bindings:
                if binding.target_group_id not in self.groups:
                    self.logger.warning(f"绑定集合 {binding.binding_id} 的目标小组 {binding.target_group_id} 不存在")
                    continue

                group = self.groups[binding.target_group_id]
                assigned_count = 0

                for member in binding.members:
                    if member.student_id not in self.placed_student_ids:
                        if group.get_remaining_capacity() > 0:
                            group.add_member(member)
                            self.placed_student_ids.add(member.student_id)
                            assigned_count += 1

                self.logger.info(f"直接委派绑定集合 {binding.binding_id}: {assigned_count}/{binding.get_size()} 成员分配到小组 {group.group_id}")

            # 处理自由分配的绑定集合
            for binding in free_bindings:
                # 寻找合适的小组
                suitable_groups = [g for g in self.groups.values()
                                 if g.get_remaining_capacity() >= binding.get_size()]

                if not suitable_groups:
                    self.logger.warning(f"绑定集合 {binding.binding_id} 无法找到合适的小组")
                    continue

                # 选择剩余空位最多的小组
                best_group = max(suitable_groups, key=lambda g: g.get_remaining_capacity())

                assigned_count = 0
                for member in binding.members:
                    if member.student_id not in self.placed_student_ids:
                        best_group.add_member(member)
                        self.placed_student_ids.add(member.student_id)
                        assigned_count += 1

                self.logger.info(f"自由分配绑定集合 {binding.binding_id}: {assigned_count}/{binding.get_size()} 成员分配到小组 {best_group.group_id}")

            self.logger.info("绑定集合分配完成")
            return True

        except Exception as e:
            self.logger.error(f"绑定集合分配失败: {str(e)}")
            return False

    def _assign_direct_individuals(self) -> bool:
        """阶段三：直接委派的落单个人"""
        try:
            self.logger.info("阶段三：开始分配直接委派的落单个人...")

            assigned_count = 0
            for student_id, target_group_id in self.direct_assignments.items():
                if student_id in self.placed_student_ids:
                    continue  # 已经分配过

                if student_id not in self.volunteers:
                    self.logger.warning(f"直接委派志愿者 {student_id} 不存在")
                    continue

                if target_group_id not in self.groups:
                    self.logger.warning(f"直接委派目标小组 {target_group_id} 不存在")
                    continue

                volunteer = self.volunteers[student_id]
                group = self.groups[target_group_id]

                if group.get_remaining_capacity() > 0:
                    group.add_member(volunteer)
                    self.placed_student_ids.add(student_id)
                    assigned_count += 1
                else:
                    self.logger.warning(f"小组 {target_group_id} 已满，无法分配志愿者 {student_id}")

            self.logger.info(f"直接委派个人分配完成: {assigned_count} 个")
            return True

        except Exception as e:
            self.logger.error(f"直接委派个人分配失败: {str(e)}")
            return False

    def _assign_remaining_internal_family(self) -> bool:
        """阶段四：剩余内部与家属志愿者"""
        try:
            self.logger.info("阶段四：开始分配剩余内部与家属志愿者...")

            # 获取未分配的内部和家属志愿者
            remaining_volunteers = [v for v in self.volunteers.values()
                                  if v.student_id not in self.placed_student_ids and
                                  v.volunteer_type in [VolunteerType.INTERNAL, VolunteerType.FAMILY]]

            # 按小组容量排序，优先分配到容量大的小组
            available_groups = sorted([(g, g.get_remaining_capacity()) for g in self.groups.values() if g.get_remaining_capacity() > 0],
                                    key=lambda x: x[1], reverse=True)

            assigned_count = 0
            group_index = 0

            for volunteer in remaining_volunteers:
                # 找到有空位的小组
                while group_index < len(available_groups) and available_groups[group_index][1] <= 0:
                    group_index += 1

                if group_index >= len(available_groups):
                    break

                group, _ = available_groups[group_index]
                group.add_member(volunteer)
                self.placed_student_ids.add(volunteer.student_id)
                assigned_count += 1

                # 更新剩余容量
                available_groups[group_index] = (group, group.get_remaining_capacity())

            self.logger.info(f"剩余内部/家属志愿者分配完成: {assigned_count} 个")
            return True

        except Exception as e:
            self.logger.error(f"剩余内部/家属志愿者分配失败: {str(e)}")
            return False

    def _assign_special_roles(self) -> bool:
        """阶段五：特殊身份填充"""
        try:
            self.logger.info("阶段五：开始分配特殊身份...")

            # 更新小组状态
            for group in self.groups.values():
                group.has_lightning = any(m.has_special_role(SpecialRole.LIGHTNING) for m in group.members)
                group.has_photography = any(m.has_special_role(SpecialRole.PHOTOGRAPHY) for m in group.members)

            # 分配小闪电
            if not self._assign_lightning_volunteers():
                return False

            # 分配摄影志愿者
            if not self._assign_photography_volunteers():
                return False

            self.logger.info("特殊身份分配完成")
            return True

        except Exception as e:
            self.logger.error(f"特殊身份分配失败: {str(e)}")
            return False

    def _assign_lightning_volunteers(self) -> bool:
        """分配小闪电志愿者"""
        try:
            # 获取需要小闪电的小组
            needy_groups = [g for g in self.groups.values()
                          if not g.has_lightning and g.get_remaining_capacity() > 0]

            if not needy_groups:
                return True

            # 获取有资格的小闪电候选人（未分配的普通志愿者）
            candidates = [v for v in self.volunteers.values()
                         if v.student_id not in self.placed_student_ids and
                         v.is_eligible_for_lightning()]

            # 按小闪电成绩降序排序
            candidates.sort(key=lambda v: v.lightning_score, reverse=True)

            assigned_count = 0
            for group in needy_groups:
                if not candidates:
                    break

                best_candidate = candidates[0]
                group.add_member(best_candidate)
                self.placed_student_ids.add(best_candidate.student_id)
                candidates.pop(0)
                assigned_count += 1

            self.logger.info(f"小闪电志愿者分配完成: {assigned_count} 个")

            # 输出小闪电分配详情
            self._output_lightning_assignment_details()

            return True

        except Exception as e:
            self.logger.error(f"小闪电志愿者分配失败: {str(e)}")
            return False

    def _assign_photography_volunteers(self) -> bool:
        """分配摄影志愿者"""
        try:
            # 获取需要摄影志愿者的小组
            needy_groups = [g for g in self.groups.values()
                          if not g.has_photography and g.get_remaining_capacity() > 0]

            if not needy_groups:
                return True

            # 获取有资格的摄影候选人（未分配的普通志愿者）
            candidates = [v for v in self.volunteers.values()
                         if v.student_id not in self.placed_student_ids and
                         v.is_eligible_for_photography()]

            # 按摄影成绩降序排序
            candidates.sort(key=lambda v: v.photography_score, reverse=True)

            assigned_count = 0
            for group in needy_groups:
                if not candidates:
                    break

                best_candidate = candidates[0]
                group.add_member(best_candidate)
                self.placed_student_ids.add(best_candidate.student_id)
                candidates.pop(0)
                assigned_count += 1

            self.logger.info(f"摄影志愿者分配完成: {assigned_count} 个")

            # 输出摄影分配详情
            self._output_photography_assignment_details()

            return True

        except Exception as e:
            self.logger.error(f"摄影志愿者分配失败: {str(e)}")
            return False

    def _assign_lightning_volunteers_properly(self) -> bool:
        """按照文档要求正确分配小闪电志愿者"""
        try:
            self.logger.info("阶段四：开始正确分配小闪电志愿者...")

            # 1. 构建小闪电候选人集合（从统一面试打分表中）
            lightning_candidates = []

            for volunteer in self.volunteers.values():
                if volunteer.is_eligible_for_lightning():
                    lightning_candidates.append({
                        'student_id': volunteer.student_id,
                        'name': volunteer.name,
                        'score': volunteer.lightning_score,
                        'volunteer': volunteer
                    })

            # 2. 按得分大小降序排列
            lightning_candidates.sort(key=lambda x: x['score'], reverse=True)

            self.logger.info(f"找到 {len(lightning_candidates)} 个小闪电候选人")

            # 3. 为每个还没有小闪电的小组分配小闪电
            groups_without_lightning = []
            for group_id, group in self.groups.items():
                if not group.has_lightning and group.get_remaining_capacity() > 0:
                    groups_without_lightning.append(group)

            self.logger.info(f"需要小闪电的小组: {len(groups_without_lightning)} 个")

            assigned_count = 0
            failed_assignments = []

            # 4. 按文档描述的逻辑进行分配
            for group in groups_without_lightning:
                while lightning_candidates and group.get_remaining_capacity() > 0:
                    # 取出得分最高的人
                    candidate = lightning_candidates[0]
                    student_id = candidate['student_id']

                    # 检查志愿者是否已经在其他小组中
                    if student_id in self.placed_student_ids:
                        # 已经在其他小组中，无法分配为小闪电
                        self.logger.warning(f"小闪电候选人 {candidate['name']} ({student_id}) 已在其他小组，跳过分配")
                        lightning_candidates.pop(0)
                        continue

                    # 检查志愿者是否是正式普通志愿者
                    if candidate['volunteer'].volunteer_type != VolunteerType.NORMAL:
                        self.logger.warning(f"小闪电候选人 {candidate['name']} ({student_id}) 不是正式普通志愿者，跳过分配")
                        lightning_candidates.pop(0)
                        continue

                    # 分配小闪电到该小组
                    candidate['volunteer'].add_special_role(SpecialRole.LIGHTNING)
                    group.add_member(candidate['volunteer'])
                    self.placed_student_ids.add(student_id)

                    # 记录分配成功
                    assigned_count += 1
                    self.logger.info(f"小闪电 {candidate['name']} ({student_id}, 成绩: {candidate['score']}) 分配到小组 {group.group_id}")

                    # 从候选人集合中删除
                    lightning_candidates.pop(0)
                    break  # 每个小组只分配一个小闪电
                else:
                    # 没有更多候选人或小组已满
                    if group.get_remaining_capacity() > 0:
                        failed_assignments.append(f"小组 {group.group_id} (剩余容量: {group.get_remaining_capacity()})")

            # 5. 记录分配结果
            self.logger.info(f"小闪电分配完成: {assigned_count} 个小组成功分配")
            if lightning_candidates:
                self.logger.info(f"剩余未分配的小闪电候选人: {len(lightning_candidates)} 个")
                for candidate in lightning_candidates[:5]:  # 显示前5个
                    self.logger.info(f"  - {candidate['name']} ({candidate['student_id']}, 成绩: {candidate['score']})")

            if failed_assignments:
                self.logger.warning(f"未能分配小闪电的小组: {len(failed_assignments)} 个")
                for reason in failed_assignments:
                    self.logger.warning(f"  - {reason}")

            return True

        except Exception as e:
            self.logger.error(f"小闪电分配失败: {str(e)}")
            return False

    def _assign_photography_volunteers_properly(self) -> bool:
        """按照文档要求正确分配摄影志愿者"""
        try:
            self.logger.info("阶段五：开始正确分配摄影志愿者...")

            # 1. 构建摄影候选人集合（从统一面试打分表中）
            photography_candidates = []

            for volunteer in self.volunteers.values():
                if volunteer.is_eligible_for_photography():
                    photography_candidates.append({
                        'student_id': volunteer.student_id,
                        'name': volunteer.name,
                        'score': volunteer.photography_score,
                        'volunteer': volunteer
                    })

            # 2. 按得分大小降序排列
            photography_candidates.sort(key=lambda x: x['score'], reverse=True)

            self.logger.info(f"找到 {len(photography_candidates)} 个摄影候选人")

            # 3. 为每个还没有摄影志愿者的小组分配摄影志愿者
            groups_without_photography = []
            for group_id, group in self.groups.items():
                if not group.has_photography and group.get_remaining_capacity() > 0:
                    groups_without_photography.append(group)

            self.logger.info(f"需要摄影志愿者的小组: {len(groups_without_photography)} 个")

            assigned_count = 0
            failed_assignments = []

            # 4. 按文档描述的逻辑进行分配
            for group in groups_without_photography:
                while photography_candidates and group.get_remaining_capacity() > 0:
                    # 取出得分最高的人
                    candidate = photography_candidates[0]
                    student_id = candidate['student_id']

                    # 检查志愿者是否已经在其他小组中
                    if student_id in self.placed_student_ids:
                        # 已经在其他小组中，无法分配为摄影志愿者
                        self.logger.warning(f"摄影候选人 {candidate['name']} ({student_id}) 已在其他小组，跳过分配")
                        photography_candidates.pop(0)
                        continue

                    # 检查志愿者是否是正式普通志愿者
                    if candidate['volunteer'].volunteer_type != VolunteerType.NORMAL:
                        self.logger.warning(f"摄影候选人 {candidate['name']} ({student_id}) 不是正式普通志愿者，跳过分配")
                        photography_candidates.pop(0)
                        continue

                    # 分配摄影志愿者到该小组
                    candidate['volunteer'].add_special_role(SpecialRole.PHOTOGRAPHY)
                    group.add_member(candidate['volunteer'])
                    self.placed_student_ids.add(student_id)

                    # 记录分配成功
                    assigned_count += 1
                    self.logger.info(f"摄影志愿者 {candidate['name']} ({student_id}, 成绩: {candidate['score']}) 分配到小组 {group.group_id}")

                    # 从候选人集合中删除
                    photography_candidates.pop(0)
                    break  # 每个小组只分配一个摄影志愿者
                else:
                    # 没有更多候选人或小组已满
                    if group.get_remaining_capacity() > 0:
                        failed_assignments.append(f"小组 {group.group_id} (剩余容量: {group.get_remaining_capacity()})")

            # 5. 记录分配结果
            self.logger.info(f"摄影志愿者分配完成: {assigned_count} 个小组成功分配")
            if photography_candidates:
                self.logger.info(f"剩余未分配的摄影候选人: {len(photography_candidates)} 个")
                for candidate in photography_candidates[:5]:  # 显示前5个
                    self.logger.info(f"  - {candidate['name']} ({candidate['student_id']}, 成绩: {candidate['score']})")

            if failed_assignments:
                self.logger.warning(f"未能分配摄影志愿者的小组: {len(failed_assignments)} 个")
                for reason in failed_assignments:
                    self.logger.warning(f"  - {reason}")

            return True

        except Exception as e:
            self.logger.error(f"摄影志愿者分配失败: {str(e)}")
            return False

    def _final_fill(self) -> bool:
        """阶段六：最终填充"""
        try:
            self.logger.info("阶段六：开始最终填充...")

            # 获取所有未分配的正式普通志愿者
            remaining_volunteers = [v for v in self.volunteers.values()
                                  if v.student_id not in self.placed_student_ids]

            # 按归一化成绩降序排序（成绩高的优先分配）
            remaining_volunteers.sort(key=lambda v: v.normalized_score or 0, reverse=True)

            # 获取所有有空位的小组
            available_spots = []
            for group in self.groups.values():
                remaining = group.get_remaining_capacity()
                for _ in range(remaining):
                    available_spots.append(group)

            assigned_count = 0
            for i, volunteer in enumerate(remaining_volunteers):
                if i >= len(available_spots):
                    self.logger.warning(f"没有足够的空位分配志愿者 {volunteer.name}")
                    break

                group = available_spots[i]
                group.add_member(volunteer)
                self.placed_student_ids.add(volunteer.student_id)
                assigned_count += 1

            self.logger.info(f"最终填充完成: {assigned_count} 个")
            return True

        except Exception as e:
            self.logger.error(f"最终填充失败: {str(e)}")
            return False

    def _output_lightning_assignment_details(self) -> None:
        """输出小闪电分配详情"""
        try:
            self.logger.info("=" * 50)
            self.logger.info("小闪电志愿者分配详情")
            self.logger.info("=" * 50)

            assigned_lightning = []
            missing_lightning = []

            for group_id, group in sorted(self.groups.items()):
                lightning_members = group.get_members_with_role(SpecialRole.LIGHTNING)
                if lightning_members:
                    lightning_vol = lightning_members[0]
                    assigned_lightning.append({
                        'group_id': group_id,
                        'position': group.position.name,
                        'name': lightning_vol.name,
                        'student_id': lightning_vol.student_id,
                        'score': lightning_vol.lightning_score or 0
                    })
                    self.logger.info(f"小组 {group_id} ({group.position.name}): {lightning_vol.name} ({lightning_vol.student_id}) - 成绩: {lightning_vol.lightning_score}")
                else:
                    missing_lightning.append(group_id)
                    self.logger.warning(f"小组 {group_id} ({group.position.name}): 未分配小闪电志愿者")

            self.logger.info("=" * 50)
            self.logger.info(f"小闪电分配统计: 已分配 {len(assigned_lightning)} 个小组，缺失 {len(missing_lightning)} 个小组")

            # 保存到报告文件
            report_path = os.path.join(os.path.dirname(get_file_path('master_schedule')), "小闪电分配报告.txt")
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("小闪电志愿者分配报告\n")
                f.write("=" * 50 + "\n\n")

                f.write(f"分配时间: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"总小组数: {len(self.groups)}\n")
                f.write(f"已分配小闪电: {len(assigned_lightning)} 个小组\n")
                f.write(f"缺失小闪电: {len(missing_lightning)} 个小组\n\n")

                f.write("详细分配情况:\n")
                f.write("-" * 50 + "\n")
                for item in sorted(assigned_lightning, key=lambda x: x['group_id']):
                    f.write(f"小组 {item['group_id']} ({item['position']}):\n")
                    f.write(f"  姓名: {item['name']}\n")
                    f.write(f"  学号: {item['student_id']}\n")
                    f.write(f"  成绩: {item['score']}\n\n")

                if missing_lightning:
                    f.write("未分配小闪电的小组:\n")
                    f.write("-" * 30 + "\n")
                    for group_id in sorted(missing_lightning):
                        group = self.groups[group_id]
                        f.write(f"小组 {group_id} ({group.position.name})\n")

            self.logger.info(f"小闪电分配报告已保存到: {report_path}")

        except Exception as e:
            self.logger.error(f"输出小闪电分配详情失败: {str(e)}")

    def _output_photography_assignment_details(self) -> None:
        """输出摄影分配详情"""
        try:
            self.logger.info("=" * 50)
            self.logger.info("摄影志愿者分配详情")
            self.logger.info("=" * 50)

            assigned_photography = []
            missing_photography = []

            for group_id, group in sorted(self.groups.items()):
                photography_members = group.get_members_with_role(SpecialRole.PHOTOGRAPHY)
                if photography_members:
                    photography_vol = photography_members[0]
                    assigned_photography.append({
                        'group_id': group_id,
                        'position': group.position.name,
                        'name': photography_vol.name,
                        'student_id': photography_vol.student_id,
                        'score': photography_vol.photography_score or 0
                    })
                    self.logger.info(f"小组 {group_id} ({group.position.name}): {photography_vol.name} ({photography_vol.student_id}) - 成绩: {photography_vol.photography_score}")
                else:
                    missing_photography.append(group_id)
                    self.logger.warning(f"小组 {group_id} ({group.position.name}): 未分配摄影志愿者")

            self.logger.info("=" * 50)
            self.logger.info(f"摄影分配统计: 已分配 {len(assigned_photography)} 个小组，缺失 {len(missing_photography)} 个小组")

            # 保存到报告文件
            report_path = os.path.join(os.path.dirname(get_file_path('master_schedule')), "摄影分配报告.txt")
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("摄影志愿者分配报告\n")
                f.write("=" * 50 + "\n\n")

                f.write(f"分配时间: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"总小组数: {len(self.groups)}\n")
                f.write(f"已分配摄影: {len(assigned_photography)} 个小组\n")
                f.write(f"缺失摄影: {len(missing_photography)} 个小组\n\n")

                f.write("详细分配情况:\n")
                f.write("-" * 50 + "\n")
                for item in sorted(assigned_photography, key=lambda x: x['group_id']):
                    f.write(f"小组 {item['group_id']} ({item['position']}):\n")
                    f.write(f"  姓名: {item['name']}\n")
                    f.write(f"  学号: {item['student_id']}\n")
                    f.write(f"  成绩: {item['score']}\n\n")

                if missing_photography:
                    f.write("未分配摄影的小组:\n")
                    f.write("-" * 30 + "\n")
                    for group_id in sorted(missing_photography):
                        group = self.groups[group_id]
                        f.write(f"小组 {group_id} ({group.position.name})\n")

            self.logger.info(f"摄影分配报告已保存到: {report_path}")

        except Exception as e:
            self.logger.error(f"输出摄影分配详情失败: {str(e)}")

    def _generate_special_roles_summary(self) -> None:
        """生成特殊身份分配汇总报告"""
        try:
            self.logger.info("生成特殊身份分配汇总报告...")

            # 生成综合报告
            report_path = os.path.join(os.path.dirname(get_file_path('master_schedule')), "特殊身份分配汇总.txt")
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("特殊身份分配汇总报告\n")
                f.write("=" * 60 + "\n\n")

                f.write(f"生成时间: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"总小组数: {len(self.groups)}\n\n")

                # 组长统计
                groups_with_leaders = sum(1 for g in self.groups.values() if g.leader is not None)
                f.write(f"组长分配: {groups_with_leaders}/{len(self.groups)} 个小组\n\n")

                # 小闪电统计
                lightning_count = sum(1 for g in self.groups.values() if g.has_lightning)
                f.write(f"小闪电分配: {lightning_count}/{len(self.groups)} 个小组\n\n")

                # 摄影统计
                photography_count = sum(1 for g in self.groups.values() if g.has_photography)
                f.write(f"摄影分配: {photography_count}/{len(self.groups)} 个小组\n\n")

                # 完整分配统计
                complete_groups = sum(1 for g in self.groups.values()
                                    if g.leader is not None and g.has_lightning and g.has_photography)
                f.write(f"完整分配(组长+小闪电+摄影): {complete_groups}/{len(self.groups)} 个小组\n\n")

                # 详细小组列表
                f.write("小组详细分配情况:\n")
                f.write("-" * 60 + "\n")
                for group_id, group in sorted(self.groups.items()):
                    status_parts = []
                    if group.leader:
                        status_parts.append(f"组长:{group.leader.name}")
                    if group.has_lightning:
                        lightning_vol = group.get_members_with_role(SpecialRole.LIGHTNING)[0]
                        status_parts.append(f"小闪电:{lightning_vol.name}")
                    if group.has_photography:
                        photography_vol = group.get_members_with_role(SpecialRole.PHOTOGRAPHY)[0]
                        status_parts.append(f"摄影:{photography_vol.name}")

                    status = " | ".join(status_parts) if status_parts else "未分配特殊角色"
                    f.write(f"小组 {group_id:2d} ({group.position.name:12s}): {status}\n")

            self.logger.info(f"特殊身份分配汇总报告已保存到: {report_path}")

        except Exception as e:
            self.logger.error(f"生成特殊身份分配汇总报告失败: {str(e)}")

    def generate_output(self) -> bool:
        """生成输出文件"""
        try:
            self.logger.info("开始生成输出文件...")

            output_path = get_file_path('master_schedule')
            output_dir = os.path.dirname(output_path)
            os.makedirs(output_dir, exist_ok=True)

            # 构建数据行
            data_rows = []

            for group in sorted(self.groups.values(), key=lambda g: g.group_id):
                # 组长置顶
                leader = group.leader
                if leader:
                    leader_row = self._create_volunteer_row(group, leader, True)
                    data_rows.append(leader_row)

                # 其他成员（按绑定集合相邻排列）
                other_members = [m for m in group.members if m != leader]
                sorted_members = self._sort_members_by_binding(other_members)

                for member in sorted_members:
                    member_row = self._create_volunteer_row(group, member, False)
                    data_rows.append(member_row)

            # 创建DataFrame
            columns = [
                '小组号', '岗位名称', '岗位简介', '小组长',
                '学号', '姓名', '姓名拼音', '性别', '证件类型', '证件号',
                '出生日期', '学院', '身高', '邮箱', '手机号', 'QQ号',
                '微信号', '政治面貌', '第几次做马拉松志愿者', '校区', '宿舍楼栋',
                '衣服尺码', '_bg_color'
            ]

            df = pd.DataFrame(data_rows, columns=columns)

            # 确保学号、证件号、手机号、QQ号为字符串格式
            str_columns = ['学号', '证件号', '手机号', 'QQ号']
            for col in str_columns:
                if col in df.columns:
                    df[col] = df[col].astype(str)

            # 使用openpyxl保存Excel并设置背景色
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import PatternFill, Font, Alignment

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "总表"

            # 将DataFrame写入工作表
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # 设置表头格式
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")  # 深蓝色
            header_font = Font(bold=True, color="FFFFFF")  # 白色粗体
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # 第一行是表头
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # 为每一行设置背景色和格式（跳过表头）
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                # 获取最后一列的颜色值（_bg_color列）
                color_value = row[-1].value  # _bg_color列的值
                if color_value and color_value != "FFFFFF":  # 跳过白色
                    # 转换为openpyxl颜色格式
                    color = f"FF{color_value}" if len(color_value) == 6 else color_value
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

                    # 为整行应用背景色（除了最后一列_bg_color）
                    for col_idx, cell in enumerate(row[:-1]):  # 跳过最后一列
                        cell.fill = fill

                # 为所有数据行设置居中对齐
                for cell in row[:-1]:  # 跳过最后一列
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # 删除_bg_color列（这一列只是临时存储颜色信息）
            ws.delete_cols(ws.max_column)

            # 设置列宽（优化显示）
            column_widths = {
                'A': 10,   # 小组号
                'B': 25,   # 岗位名称
                'C': 40,   # 岗位简介
                'D': 15,   # 小组长
                'E': 20,   # 学号
                'F': 12,   # 姓名
                'G': 20,   # 姓名拼音
                'H': 8,    # 性别
                'I': 10,   # 证件类型
                'J': 25,   # 证件号
                'K': 15,   # 出生日期
                'L': 25,   # 学院
                'M': 10,   # 身高
                'N': 30,   # 邮箱
                'O': 15,   # 手机号
                'P': 15,   # QQ号
                'Q': 20,   # 微信号
                'R': 15,   # 政治面貌
                'S': 20,   # 第几次做马拉松志愿者
                'T': 15,   # 校区
                'U': 15,   # 宿舍楼栋
                'V': 10,   # 衣服尺码
            }

            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # 冻结首行
            ws.freeze_panes = 'A2'

            # 保存文件
            wb.save(output_path)

            self.logger.info(f"输出文件生成成功: {output_path}")
            self.logger.info(f"总行数: {len(df)} (不含表头)")
            self.logger.info(f"已为所有志愿者行设置背景色")

            # 生成特殊身份分配汇总报告
            self._generate_special_roles_summary()

            return True

        except Exception as e:
            self.logger.error(f"输出文件生成失败: {str(e)}")
            return False

    def _create_volunteer_row(self, group: Group, volunteer: Volunteer, is_leader: bool = False) -> List[str]:
        """创建志愿者数据行"""
        row = [
            group.group_id,
            group.position.name,
            group.position.description,
            group.leader.name if group.leader else "",
            volunteer.student_id,
            volunteer.name,
            volunteer.name_pinyin or "",
            volunteer.gender or "",
            volunteer.id_type or "",
            volunteer.id_number or "",
            volunteer.birth_date or "",
            volunteer.college or "",
            volunteer.height or "",
            volunteer.email or "",
            volunteer.phone or "",
            volunteer.qq or "",
            volunteer.wechat or "",
            volunteer.political_status or "",
            volunteer.marathon_count or "",
            volunteer.campus or "",
            volunteer.dorm_building or "",
            volunteer.clothes_size or "",
            self._calculate_background_color(volunteer)
        ]

        return row

    def save_metadata(self) -> bool:
        """保存更新后的元数据文件，保持原有结构，只添加团体颜色映射"""
        try:
            metadata_path = get_file_path('metadata')
            os.makedirs(os.path.dirname(metadata_path), exist_ok=True)

            # 读取现有的metadata.json文件
            existing_metadata = {}
            if os.path.exists(metadata_path):
                with open(metadata_path, 'r', encoding='utf-8') as f:
                    existing_metadata = json.load(f)
                self.logger.info(f"读取现有metadata文件，包含 {len(existing_metadata)} 个顶级键")

            # 添加团体颜色映射到现有结构中
            if hasattr(self.metadata, 'group_color_mapping') and self.metadata.group_color_mapping:
                existing_metadata['group_color_mapping'] = self.metadata.group_color_mapping
                self.logger.info(f"添加团体颜色映射: {len(self.metadata.group_color_mapping)} 个团体")
            else:
                existing_metadata['group_color_mapping'] = {}

            # 确保JSON序列化兼容
            def convert_numpy_types(obj):
                if hasattr(obj, 'tolist'):
                    return obj.tolist()
                elif hasattr(obj, 'item'):
                    return obj.item()
                elif isinstance(obj, dict):
                    return {k: convert_numpy_types(v) for k, v in obj.items()}
                elif isinstance(obj, list):
                    return [convert_numpy_types(item) for item in obj]
                return obj

            metadata_serializable = convert_numpy_types(existing_metadata)

            with open(metadata_path, 'w', encoding='utf-8') as f:
                json.dump(metadata_serializable, f, ensure_ascii=False, indent=2)

            self.logger.info(f"元数据文件已更新保存: {metadata_path}")
            return True

        except Exception as e:
            self.logger.error(f"保存元数据文件失败: {str(e)}")
            return False

    def validate_result(self) -> bool:
        """验证分配结果"""
        try:
            self.logger.info("开始验证分配结果...")

            # 1. 总人数校验
            total_assigned = len(self.placed_student_ids)
            expected_total = self.metadata.total_required_volunteers

            if total_assigned != expected_total:
                self.logger.warning(f"分配人数不匹配: 实际 {total_assigned}, 预期 {expected_total}")

            # 2. ID 唯一性校验
            assigned_ids = list(self.placed_student_ids)
            if len(assigned_ids) != len(set(assigned_ids)):
                self.logger.error("存在重复的学号分配")
                return False

            # 3. 小组完整性校验
            for group_id, group in self.groups.items():
                actual_count = len(group.members)
                required_count = group.required_count

                if actual_count != required_count:
                    self.logger.warning(f"小组 {group_id} 人数不匹配: 实际 {actual_count}, 预期 {required_count}")

            # 4. 空值校验
            for volunteer in self.volunteers.values():
                if volunteer.student_id in self.placed_student_ids:
                    if not volunteer.student_id or not volunteer.name:
                        self.logger.error(f"志愿者 {volunteer.student_id} 信息不完整")
                        return False

            self.logger.info("分配结果验证通过")
            return True

        except Exception as e:
            self.logger.error(f"分配结果验证失败: {str(e)}")
            return False

    def _sort_members_by_binding(self, members: List[Volunteer]) -> List[Volunteer]:
        """
        按绑定集合对成员进行排序，确保同一个绑定集合的成员相邻排列

        排序规则：
        1. 同一个绑定集合的成员排在一起
        2. 绑定集合按类型和大小排序（直接委派 > 情侣 > 家庭 > 团体 > 混合）
        3. 不在任何绑定集合的成员按学号排序
        4. 每个绑定集合内部按学号排序
        """
        try:
            # 构建学号到绑定集合的映射
            student_id_to_binding = {}

            for binding_set in self.binding_sets:
                for member in binding_set.members:
                    student_id_to_binding[member.student_id] = binding_set

            # 将成员分类
            binding_members = {}  # binding_id -> [members]
            unbound_members = []  # 不在任何绑定集合的成员

            for member in members:
                if member.student_id in student_id_to_binding:
                    binding_set = student_id_to_binding[member.student_id]
                    if binding_set.binding_id not in binding_members:
                        binding_members[binding_set.binding_id] = []
                    binding_members[binding_set.binding_id].append(member)
                else:
                    unbound_members.append(member)

            # 排序绑定集合
            sorted_binding_ids = sorted(
                binding_members.keys(),
                key=lambda binding_id: self._get_binding_sort_key(
                    next(bs for bs in self.binding_sets if bs.binding_id == binding_id)
                )
            )

            # 构建最终排序结果
            sorted_members = []

            # 1. 添加各个绑定集合的成员
            for binding_id in sorted_binding_ids:
                members_in_binding = sorted(
                    binding_members[binding_id],
                    key=lambda v: v.student_id  # 绑定集合内部按学号排序
                )
                sorted_members.extend(members_in_binding)

            # 2. 添加不在任何绑定集合的成员
            unbound_members.sort(key=lambda v: v.student_id)
            sorted_members.extend(unbound_members)

            self.logger.info(f"成员排序完成: {len(sorted_binding_ids)}个绑定集合, {len(unbound_members)}个独立成员")

            return sorted_members

        except Exception as e:
            self.logger.warning(f"按绑定集合排序失败，使用默认学号排序: {str(e)}")
            # 降级处理：按学号排序
            return sorted(members, key=lambda v: v.student_id)

    def _get_binding_sort_key(self, binding_set: 'BindingSet') -> tuple:
        """
        获取绑定集合的排序键

        排序优先级：
        1. 绑定类型：direct(1) > couple(2) > family(3) > group(4) > mixed(5)
        2. 绑定集合大小：大的优先
        3. 绑定ID：确保稳定排序
        """
        # 绑定类型优先级映射
        type_priority = {
            'direct': 1,  # 直接委派优先级最高
            'couple': 2,
            'family': 3,
            'group': 4,
            'mixed': 5,
        }

        priority = type_priority.get(binding_set.binding_type, 5)
        size = len(binding_set.members)

        return (priority, -size, binding_set.binding_id)


def run_scheduler() -> bool:
    """运行排表调度器"""
    try:
        scheduler = VolunteerScheduler()

        # 加载数据
        if not scheduler.load_data():
            return False

        # 执行调度
        result = scheduler.execute_scheduling()
        if not result.success:
            scheduler.logger.error(f"排表调度失败: {result.message}")
            return False

        # 生成输出
        if not scheduler.generate_output():
            return False

        # 验证结果
        if not scheduler.validate_result():
            return False

        # 保存更新后的元数据（包含团体颜色信息）
        if not scheduler.save_metadata():
            scheduler.logger.warning("元数据保存失败，但不影响主要功能")

        print(f"* 排表调度成功完成: {result.assigned_count} 个志愿者已分配")
        return True

    except Exception as e:
        print(f"* 排表调度异常: {str(e)}")
        return False


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='志愿者排表主程序')
    parser.add_argument('--verbose', '-v', action='store_true', help='详细输出模式')

    args = parser.parse_args()

    if run_scheduler():
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()